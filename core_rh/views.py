from django.db.models import Q
from django.shortcuts import render, redirect, get_object_or_404
import zipfile
import io
import os
import base64
import requests 
import re 
from urllib.parse import unquote 
from django.contrib.staticfiles import finders
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model, update_session_auth_hash
from django.contrib.auth.views import PasswordResetView, PasswordResetDoneView
from django.contrib.auth.forms import PasswordResetForm, PasswordChangeForm
from django.urls import reverse_lazy, reverse
from django.utils import timezone 
from django.http import HttpResponse 
from datetime import date, time, datetime, timedelta 
from calendar import monthrange, monthcalendar 
from django.template.loader import render_to_string 
from django.conf import settings
from django.contrib import messages
from django.core.files.base import ContentFile
from django.db.models import Sum
from django.db import transaction
# PDF e Relatórios
import pdfplumber
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors 
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle
from collections import defaultdict
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from django.views.decorators.http import require_POST
# Utils Extras
import holidays 
try:
    from weasyprint import HTML, CSS
except ImportError:
    pass

try:
    from pdf2image import convert_from_path
    HAS_PDF_CONVERTER = True
except ImportError:
    HAS_PDF_CONVERTER = False

try:
    from pypdf import PdfReader, PdfWriter
except ImportError:
    PdfReader = None
    PdfWriter = None

# Models e Forms
from .models import (
    RegistroPonto, Funcionario, Equipe, Contracheque, Ferias, 
    Atestado, ControleKM, TrechoKM, DespesaDiversa
)
from .forms import AtestadoForm, CpfPasswordResetForm

User = get_user_model()

def usuario_eh_rh(user):
    """
    Retorna True se o usuário for Superuser, ou estiver na equipe 'RH' 
    (seja como principal ou secundária).
    """
    if not user.is_authenticated: return False
    if user.is_superuser: return True
    
    # Verifica grupos do Django Admin (legado)
    if user.groups.filter(name='RH').exists(): return True

    try:
        func = user.funcionario
        # Verifica Equipe Principal (Nome exato 'RH' ou 'Recursos Humanos')
        if func.equipe and func.equipe.nome in ['RH', 'Recursos Humanos', 'Gestão de Pessoas']:
            return True
        
        # Verifica Equipes Secundárias
        if func.outras_equipes.filter(nome__in=['RH', 'Recursos Humanos', 'Gestão de Pessoas']).exists():
            return True
            
    except AttributeError:
        # Usuário sem funcionário vinculado
        pass
        
    return False

DIAS_SEMANA_PT = {
    0: 'Segunda-feira', 1: 'Terça-feira', 2: 'Quarta-feira', 3: 'Quinta-feira',
    4: 'Sexta-feira', 5: 'Sábado', 6: 'Domingo'
}

MESES_PT = {
    1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
    5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
    9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
}

def get_competencia_atual():
    hoje = timezone.now()
    if hoje.day >= 16:
        if hoje.month == 12:
            return 1, hoje.year + 1
        else:
            return hoje.month + 1, hoje.year
    else:
        return hoje.month, hoje.year

def get_competencia_anterior(mes, ano):
    if mes == 1:
        return 12, ano - 1
    return mes - 1, ano

def get_datas_competencia(mes_referencia, ano_referencia):
    if mes_referencia == 1:
        mes_anterior = 12
        ano_anterior = ano_referencia - 1
    else:
        mes_anterior = mes_referencia - 1
        ano_anterior = ano_referencia
        
    data_inicio = date(ano_anterior, mes_anterior, 16)
    data_fim = date(ano_referencia, mes_referencia, 15)
    
    return data_inicio, data_fim

def calcular_horas_trabalhadas(entrada_1_str, saida_1_str, entrada_2_str, saida_2_str):
    total = timedelta()
    try:
        if entrada_1_str and saida_1_str:
            e1 = datetime.strptime(entrada_1_str, '%H:%M')
            s1 = datetime.strptime(saida_1_str, '%H:%M')
            if s1 > e1:
                total += s1 - e1
        if entrada_2_str and saida_2_str:
            e2 = datetime.strptime(entrada_2_str, '%H:%M')
            s2 = datetime.strptime(saida_2_str, '%H:%M')
            if s2 > e2:
                total += s2 - e2
    except ValueError:
        return timedelta()
    return total

def usuario_eh_campo(user):
    if not user.is_authenticated: return False
    if user.is_superuser: return True
    try:
        func = user.funcionario
        termo = "Campo"
        if func.equipe and termo in func.equipe.nome: return True
        if func.outras_equipes.filter(nome__icontains=termo).exists(): return True
    except:
        pass
    return False

@login_required 
def home(request):
    is_gestor = False
    tem_ferias = False 
    is_campo = False
    equipes_gestor = []

    try:
        funcionario = Funcionario.objects.get(usuario=request.user)
        
        # Verifica se é gestor (Principal ou Lista) de equipes NÃO ocultas
        equipes_lideradas = Equipe.objects.filter(
            Q(oculta=False) & (Q(gestor=funcionario) | Q(gestores=funcionario))
        ).distinct()

        if equipes_lideradas.exists():
            is_gestor = True
            equipes_gestor = equipes_lideradas

        # --- LÓGICA DE FÉRIAS ATUALIZADA ---
        # O botão aparece se houver um registro de férias com arquivo gerado 
        # E se a data de fim das férias ainda não passou (data_fim >= hoje).
        hoje = timezone.now().date()
        if Ferias.objects.filter(
            funcionario=funcionario, 
            data_fim__gte=hoje
        ).exclude(arquivo_aviso='').exists():
            tem_ferias = True
        # -----------------------------------
            
        is_campo = usuario_eh_campo(request.user)
        
    except Funcionario.DoesNotExist: 
        pass 
    
    can_access_rh_area = usuario_eh_rh(request.user)
    
    return render(request, 'core_rh/index.html', {
        'is_gestor': is_gestor or request.user.is_superuser, 
        'equipes_lideradas': equipes_gestor,
        'can_access_rh_area': can_access_rh_area,
        'tem_ferias': tem_ferias,
        'is_campo': is_campo, 
    })

@login_required
def salvar_ponto_view(request):
    if request.method != 'POST':
        return redirect('folha_ponto')

    try:
        mes = int(request.POST.get('mes'))
        ano = int(request.POST.get('ano'))
    except (ValueError, TypeError):
        return redirect('folha_ponto')
        
    redirect_url = reverse_lazy('folha_ponto') + f"?mes={mes}&ano={ano}"

    try:
        funcionario = Funcionario.objects.get(usuario=request.user)
    except Funcionario.DoesNotExist:
        messages.error(request, "Perfil de funcionário não encontrado.")
        return redirect('folha_ponto')

    data_inicio, data_fim = get_datas_competencia(mes, ano)

    if RegistroPonto.objects.filter(funcionario=funcionario, data__range=[data_inicio, data_fim], assinado_gestor=True).exists():
        messages.error(request, "ERRO: Esta folha já foi fechada e assinada pelo gestor. Solicite o desbloqueio ao RH.")
        return redirect(redirect_url)

    if request.FILES.get('pdf_assinado'):
        arquivo = request.FILES['pdf_assinado']
        nome_limpo = funcionario.nome_completo.strip().replace(' ', '_')
        arquivo.name = f"Folha_{nome_limpo}_{mes:02d}_{ano}_Assinado_Colab.pdf"
        registros = RegistroPonto.objects.filter(funcionario=funcionario, data__range=[data_inicio, data_fim])
        if registros.exists():
            registros.update(assinado_funcionario=True, assinado_gestor=False)
            primeiro_reg = registros.first()
            primeiro_reg.arquivo_anexo = arquivo
            primeiro_reg.save()
            messages.success(request, "Documento enviado com sucesso! A assinatura do gestor foi resetada (se houver).")
        else:
             messages.error(request, "Nenhum registro de ponto encontrado para anexar o arquivo.")
        return redirect(redirect_url)

    delta_dias = (data_fim - data_inicio).days

    for i in range(delta_dias + 1):
        data_ponto = data_inicio + timedelta(days=i)
        dia_num = data_ponto.day
        entrada_1_str = request.POST.get(f'entrada_1_{dia_num}', '').strip()
        saida_1_str = request.POST.get(f'saida_1_{dia_num}', '').strip()
        entrada_2_str = request.POST.get(f'entrada_2_{dia_num}', '').strip()
        saida_2_str = request.POST.get(f'saida_2_{dia_num}', '').strip()
        entrada_extra_str = request.POST.get(f'entrada_extra_{dia_num}', '').strip()
        saida_extra_str = request.POST.get(f'saida_extra_{dia_num}', '').strip()
        observacoes = request.POST.get(f'observacoes_{dia_num}', '').strip()
        
        try:
            registro = RegistroPonto.objects.get(funcionario=funcionario, data=data_ponto)
            registro_existe = True
        except RegistroPonto.DoesNotExist:
            registro_existe = False
            
        if not any([entrada_1_str, saida_1_str, entrada_2_str, saida_2_str, entrada_extra_str, saida_extra_str, observacoes]):
            if registro_existe:
                registro.delete()
            continue
            
        try:
            entrada_1 = time.fromisoformat(entrada_1_str) if entrada_1_str else None
            saida_1 = time.fromisoformat(saida_1_str) if saida_1_str else None
            entrada_2 = time.fromisoformat(entrada_2_str) if entrada_2_str else None
            saida_2 = time.fromisoformat(saida_2_str) if saida_2_str else None
            entrada_extra = time.fromisoformat(entrada_extra_str) if entrada_extra_str else None
            saida_extra = time.fromisoformat(saida_extra_str) if saida_extra_str else None
        except ValueError:
            continue 

        RegistroPonto.objects.update_or_create(
            funcionario=funcionario, 
            data=data_ponto,
            defaults={
                'entrada_manha': entrada_1,      
                'saida_almoco': saida_1,        
                'volta_almoco': entrada_2,        
                'saida_tarde': saida_2,          
                'extra_entrada': entrada_extra,
                'extra_saida': saida_extra,      
                'observacao': observacoes,        
            }
        )

    if not request.FILES.get('pdf_assinado'):
        messages.success(request, "Dados de ponto salvos com sucesso!")          
    
    return redirect(redirect_url)

@login_required
def gerar_pdf_ponto_view(request):
    mes_atual, ano_atual = get_competencia_atual()
    try:
        mes = int(request.GET.get('mes', mes_atual))
        ano = int(request.GET.get('ano', ano_atual))
        target_func_id = request.GET.get('funcionario_id')
    except ValueError: return HttpResponse("Parâmetros inválidos.", status=400)

    if target_func_id:
        try:
            alvo = Funcionario.objects.get(id=target_func_id)
            if usuario_eh_rh(request.user) or request.user.is_superuser: 
                funcionario = alvo
            else:
                try:
                    gestor = Funcionario.objects.get(usuario=request.user)
                    
                    # --- CORREÇÃO: Permissão para gerar PDF como gestor ---
                    is_gestor_autorizado = Equipe.objects.filter(
                        Q(gestor=gestor) | Q(gestores=gestor)
                    ).filter(Q(id=alvo.equipe.id) | Q(id__in=alvo.outras_equipes.values_list('id', flat=True))).exists()

                    if is_gestor_autorizado:
                        funcionario = alvo
                    else: return HttpResponse("Acesso negado.", status=403)
                except Funcionario.DoesNotExist: return HttpResponse("Perfil não encontrado.", status=403)
        except Funcionario.DoesNotExist: return HttpResponse("Funcionário não encontrado.", status=404)
    else:
        try: funcionario = Funcionario.objects.get(usuario=request.user)
        except Funcionario.DoesNotExist: return HttpResponse("Perfil não encontrado.", status=404)

    feriados_br = holidays.BR(state='DF', years=ano) 
    if hasattr(funcionario, 'estado_sigla') and funcionario.estado_sigla:
        try: feriados_br = holidays.BR(state=funcionario.estado_sigla, years=ano)
        except: pass
            
    data_inicio, data_fim = get_datas_competencia(mes, ano)
    registros_dict = {r.data.day: r for r in RegistroPonto.objects.filter(funcionario=funcionario, data__range=[data_inicio, data_fim]).order_by('data')}
    
    atestados_periodo = Atestado.objects.filter(
        funcionario=funcionario,
        status='Aprovado',
        data_inicio__lte=data_fim
    )

    dias_do_mes = []
    total_horas_delta = timedelta()
    total_extras_delta = timedelta()
    
    delta_dias = (data_fim - data_inicio).days
    for i in range(delta_dias + 1):
        data_atual = data_inicio + timedelta(days=i)
        registro = registros_dict.get(data_atual.day)
        
        tipo_atestado = None
        for atestado in atestados_periodo:
            if atestado.tipo == 'DIAS':
                fim = atestado.data_inicio + timedelta(days=(atestado.qtd_dias or 1) - 1)
                if atestado.data_inicio <= data_atual <= fim:
                    tipo_atestado = 'DIAS'
                    break
            elif atestado.tipo == 'HORAS' and atestado.data_inicio == data_atual:
                tipo_atestado = 'HORAS'

        eh_feriado = data_atual in feriados_br
        nome_feriado = feriados_br.get(data_atual).upper() if eh_feriado else ""

        if not registro:
            class MockReg:
                entrada_manha = None
                saida_almoco = None
                volta_almoco = None
                saida_tarde = None
                extra_entrada = None
                extra_saida = None
                horas_extra = ""
                observacao = ""
            registro = MockReg()
        
        if tipo_atestado == 'DIAS':
            registro.observacao = "Atestado Médico"
        elif tipo_atestado == 'HORAS':
            if not registro.observacao: 
                registro.observacao = "Atestado de Comparecimento"

        td_normal = timedelta()
        if hasattr(registro, 'pk') and registro.entrada_manha and registro.saida_almoco:
            dt_e1 = datetime.combine(date.min, registro.entrada_manha)
            dt_s1 = datetime.combine(date.min, registro.saida_almoco)
            if dt_s1 > dt_e1: td_normal += (dt_s1 - dt_e1)
        if hasattr(registro, 'pk') and registro.volta_almoco and registro.saida_tarde:
            dt_e2 = datetime.combine(date.min, registro.volta_almoco)
            dt_s2 = datetime.combine(date.min, registro.saida_tarde)
            if dt_s2 > dt_e2: td_normal += (dt_s2 - dt_e2)
        total_horas_delta += td_normal

        td_extra = timedelta()
        if hasattr(registro, 'pk') and registro.extra_entrada and registro.extra_saida:
            dt_ex_ent = datetime.combine(date.min, registro.extra_entrada)
            dt_ex_sai = datetime.combine(date.min, registro.extra_saida)
            if dt_ex_sai > dt_ex_ent: td_extra = dt_ex_sai - dt_ex_ent
        total_extras_delta += td_extra
        
        if td_extra.total_seconds() > 0:
            registro.horas_extra = format_delta(td_extra)
        else:
            if not hasattr(registro, 'horas_extra'):
                registro.horas_extra = ""

        dias_do_mes.append({
            'data': data_atual,
            'dia_semana_nome': DIAS_SEMANA_PT[data_atual.weekday()],
            'eh_feriado': eh_feriado,
            'nome_feriado': nome_feriado,
            'registro': registro
        })

    logo_data = None
    possiveis_caminhos = [
        os.path.join(settings.BASE_DIR, 'core', 'static', 'images', 'Logo.png'),
        os.path.join(settings.BASE_DIR, 'staticfiles', 'images', 'Logo.png'),
        os.path.join(settings.BASE_DIR, 'static', 'images', 'Logo.png')
    ]
    for path in possiveis_caminhos:
        if os.path.exists(path):
            try:
                with open(path, "rb") as image_file:
                    logo_data = base64.b64encode(image_file.read()).decode('utf-8')
                break
            except Exception: pass

    context = {
        'funcionario': funcionario,
        'empresa': 'Dividata Processamento de Dados Ltda', 
        'cnpj': '20.914.172/0001-88',
        'endereco': 'Praça Governador Benedito Valadares, 84 - Sobreloja, Divinópolis - MG',
        'mes_ano': f"{mes:02d}/{ano}",
        'dias_do_mes': dias_do_mes,
        'nome_mes': f"{MESES_PT[data_inicio.month]}/{MESES_PT[mes]} {ano}",
        'total_horas': format_delta(total_horas_delta),
        'total_horas_extras': format_delta(total_extras_delta),
        'user_mock': request.user,
        'logo_b64': logo_data,
    }

    html_string = render_to_string('core_rh/pdf_folha_ponto.html', context)
    response = HttpResponse(content_type='application/pdf')
    nome_func = funcionario.nome_completo.strip().replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename="Folha_{nome_func}_{mes:02d}_{ano}.pdf"'

    try:
        from weasyprint import HTML
        HTML(string=html_string, base_url=str(settings.BASE_DIR)).write_pdf(response)
    except ImportError:
        response.write(html_string)
    
    return response

@login_required
def folha_ponto_view(request):
    mes_atual_real, ano_atual_real = get_competencia_atual()
    
    try:
        mes_solicitado = int(request.GET.get('mes', mes_atual_real))
        ano_solicitado = int(request.GET.get('ano', ano_atual_real))
    except ValueError:
        mes_solicitado = mes_atual_real
        ano_solicitado = ano_atual_real
        
    mes_anterior_permitido, ano_anterior_permitido = get_competencia_anterior(mes_atual_real, ano_atual_real)
    
    is_atual = (mes_solicitado == mes_atual_real and ano_solicitado == ano_atual_real)
    is_anterior = (mes_solicitado == mes_anterior_permitido and ano_solicitado == ano_anterior_permitido)
    
    if not (is_atual or is_anterior):
        return redirect(f'{reverse_lazy("folha_ponto")}?mes={mes_atual_real}&ano={ano_atual_real}')

    prev_mes, prev_ano, next_mes, next_ano = None, None, None, None
    if is_atual:
        prev_mes, prev_ano = mes_anterior_permitido, ano_anterior_permitido
    elif is_anterior:
        next_mes, next_ano = mes_atual_real, ano_atual_real

    funcionario = None
    feriados_br = holidays.BR(state='DF', years=ano_solicitado)
    
    data_inicio, data_fim = get_datas_competencia(mes_solicitado, ano_solicitado)
    
    atestados_periodo = []
    ferias_periodo = []

    try:
        funcionario = Funcionario.objects.get(usuario=request.user)
        if hasattr(funcionario, 'estado_sigla') and funcionario.estado_sigla:
             feriados_br = holidays.BR(state=funcionario.estado_sigla, years=ano_solicitado)
        
        atestados_periodo = Atestado.objects.filter(
            funcionario=funcionario,
            status='Aprovado',
            data_inicio__lte=data_fim
        )

        ferias_periodo = Ferias.objects.filter(
            funcionario=funcionario,
            data_inicio__lte=data_fim,
            data_fim__gte=data_inicio
        )

    except (Funcionario.DoesNotExist, NameError):
        pass 

    dias_do_mes = []
    registros_dict = {}
    is_locked = False

    if funcionario:
        registros_banco = RegistroPonto.objects.filter(
            funcionario=funcionario, 
            data__range=[data_inicio, data_fim]
        )
        registros_dict = {r.data.day: r for r in registros_banco}
        
        if registros_banco.filter(assinado_gestor=True).exists():
            is_locked = True

    delta_dias = (data_fim - data_inicio).days
    for i in range(delta_dias + 1):
        data_atual = data_inicio + timedelta(days=i)
        
        eh_fim_de_semana = data_atual.weekday() >= 5
        eh_feriado = data_atual in feriados_br
        nome_feriado = feriados_br.get(data_atual) if eh_feriado else ""
        
        eh_ferias = False
        for f in ferias_periodo:
            if f.data_inicio <= data_atual <= f.data_fim:
                eh_ferias = True
                break

        tipo_atestado = None 
        for atestado in atestados_periodo:
            if atestado.tipo == 'DIAS':
                fim_atestado = atestado.data_inicio + timedelta(days=(atestado.qtd_dias or 1) - 1)
                if atestado.data_inicio <= data_atual <= fim_atestado:
                    tipo_atestado = 'DIAS'
                    break
            elif atestado.tipo == 'HORAS':
                if atestado.data_inicio == data_atual:
                    tipo_atestado = 'HORAS'
        
        registro = registros_dict.get(data_atual.day)
        
        obs_visual = ""
        if registro and registro.observacao:
            obs_visual = registro.observacao
        elif eh_ferias:
            obs_visual = "Férias" 
        elif eh_feriado:
            obs_visual = "FERIADO"
        elif tipo_atestado == 'DIAS':
            obs_visual = "Atestado Médico"
        elif tipo_atestado == 'HORAS':
            obs_visual = "Atestado de Comparecimento"

        dias_do_mes.append({
            'data': data_atual,
            'dia_semana_nome': DIAS_SEMANA_PT[data_atual.weekday()],
            'eh_fim_de_semana': eh_fim_de_semana,
            'eh_feriado': eh_feriado,
            'nome_feriado': nome_feriado.upper() if nome_feriado else "",
            'registro': registro,
            'tipo_atestado': tipo_atestado,
            'eh_ferias': eh_ferias,
            'obs_visual': obs_visual
        })

    mes_anterior_num = data_inicio.month
    nome_mes_composto = f"{MESES_PT[mes_anterior_num]}/{MESES_PT[mes_solicitado]}"
    
    context = {
        'dias_do_mes': dias_do_mes,
        'mes_atual': mes_solicitado, 
        'ano_atual': ano_solicitado, 
        'nome_mes': nome_mes_composto,
        'funcionario': funcionario,
        'prev_mes': prev_mes,
        'prev_ano': prev_ano,
        'next_mes': next_mes,
        'next_ano': next_ano,
        'is_locked': is_locked,
    }
    
    return render(request, 'core_rh/folha_ponto.html', context)


# --- SUBSTITUA A FUNÇÃO area_gestor_view INTEIRA POR ESTA ---
@login_required
def area_gestor_view(request):
    # --- IMPORTS NECESSÁRIOS (Idealmente no topo do arquivo) ---
    from datetime import date, timedelta
    from calendar import monthcalendar
    from django.db.models import Q
    try:
        gestor = Funcionario.objects.get(usuario=request.user)
        # Pega TODAS as equipes
        todas_equipes = Equipe.objects.filter(
            Q(gestor=gestor) | Q(gestores=gestor)
        ).distinct()
    except Funcionario.DoesNotExist:
        return redirect('home')

    if not todas_equipes.exists() and not request.user.is_superuser:
        messages.error(request, "Acesso negado. Você não é gestor de nenhuma equipe.")
        return redirect('home')

    # --- SEPARAÇÃO DAS EQUIPES ---
    equipes_ponto = todas_equipes.filter(oculta=False)
    equipes_km = todas_equipes.filter(oculta=True)

    # --- LÓGICA DE DATAS (COMPETÊNCIA) ---
    try:
        mes_real, ano_real = get_competencia_atual()
    except NameError:
        hoje = date.today()
        mes_real, ano_real = hoje.month, hoje.year

    try:
        mes = int(request.GET.get('mes', mes_real))
        ano = int(request.GET.get('ano', ano_real))
    except ValueError:
        mes, ano = mes_real, ano_real

    # Navegação Meses
    if mes == 1:
        mes_ant, ano_ant = 12, ano - 1
    else:
        mes_ant, ano_ant = mes - 1, ano
        
    if mes == 12:
        mes_prox, ano_prox = 1, ano + 1
    else:
        mes_prox, ano_prox = mes + 1, ano

    nav_anterior = {'mes': mes_ant, 'ano': ano_ant}
    nav_proximo = {'mes': mes_prox, 'ano': ano_prox}

    # ==========================================
    # ABA 1: PONTO (Equipes Visíveis)
    # ==========================================
    lista_ponto = []
    if equipes_ponto.exists():
        try:
            data_inicio, data_fim = get_datas_competencia(mes, ano)
        except NameError:
            data_inicio = date(ano, mes, 1)
            import calendar
            last_day = calendar.monthrange(ano, mes)[1]
            data_fim = date(ano, mes, last_day)

        funcionarios_ponto = Funcionario.objects.filter(
            Q(equipe__in=equipes_ponto) | Q(outras_equipes__in=equipes_ponto)
        ).exclude(id=gestor.id).distinct()
        
        for func in funcionarios_ponto:
            pontos = RegistroPonto.objects.filter(funcionario=func, data__range=[data_inicio, data_fim])
            assinado_func = pontos.filter(assinado_funcionario=True).exists()
            assinado_gest = pontos.filter(assinado_gestor=True).exists()
            ponto_com_arquivo = pontos.exclude(arquivo_anexo='').first()
            url_arquivo = ponto_com_arquivo.arquivo_anexo.url if ponto_com_arquivo and ponto_com_arquivo.arquivo_anexo else None
            
            lista_ponto.append({
                'funcionario': func,
                'status_func': assinado_func,
                'status_gestor': assinado_gest,
                'pode_assinar': assinado_func and not assinado_gest,
                'arquivo_assinado_url': url_arquivo,
                'nome_download': f"Folha_{func.nome_completo}_{mes}_{ano}.pdf"
            })

    # ==========================================
    # ABA 2: KM (Equipes Ocultas)
    # ==========================================
    semanas_do_mes = []
    dados_km_semana_atual = []
    equipe_km_selecionada = None
    semana_selecionada = 1
    
    if equipes_km.exists():
        # Seleciona equipe
        km_team_id = request.GET.get('km_team')
        if km_team_id:
            equipe_km_selecionada = equipes_km.filter(id=km_team_id).first()
        if not equipe_km_selecionada:
            equipe_km_selecionada = equipes_km.first()

        # Seleciona Semana
        try: 
            semana_selecionada = int(request.GET.get('semana', 1))
        except: 
            semana_selecionada = 1

        # Calcula Semanas
        cal = monthcalendar(ano, mes)
        count_sem = 1
        range_semana_atual = None

        for week in cal:
            dias_validos = [d for d in week if d != 0]
            if not dias_validos: continue
            
            primeiro_dia = date(ano, mes, dias_validos[0])
            inicio_semana = primeiro_dia - timedelta(days=primeiro_dia.weekday())
            fim_semana = inicio_semana + timedelta(days=6)
            
            is_active = (count_sem == semana_selecionada)
            semanas_do_mes.append({
                'numero': count_sem,
                'inicio': inicio_semana,
                'fim': fim_semana,
                'active': is_active
            })
            
            if is_active:
                range_semana_atual = (inicio_semana, fim_semana)
            
            count_sem += 1

        # Fallback se não achar a semana
        if not range_semana_atual and semanas_do_mes:
            range_semana_atual = (semanas_do_mes[0]['inicio'], semanas_do_mes[0]['fim'])

        # Busca Dados da Semana
        if range_semana_atual:
            funcs_campo = Funcionario.objects.filter(
                Q(equipe=equipe_km_selecionada) | Q(outras_equipes=equipe_km_selecionada)
            ).distinct().order_by('nome_completo')

            ini, fim = range_semana_atual
            
            for f in funcs_campo:
                kms = ControleKM.objects.filter(funcionario=f, data__range=[ini, fim])
                despesas = DespesaDiversa.objects.filter(funcionario=f, data__range=[ini, fim])
                
                total_km_val = sum(k.total_km for k in kms)
                total_despesas_val = sum(d.valor for d in despesas)
                
                tem_algo = kms.exists() or despesas.exists()
                status_geral = "Vazio"
                ids_km = []
                
                valor_total_financeiro = 0.0

                if tem_algo:
                    # Cálculo Financeiro
                    fator = float(f.valor_km) if f.valor_km and f.valor_km > 0 else 1.20
                    valor_total_financeiro = (float(total_km_val) * fator) + float(total_despesas_val)

                    # Status
                    st_km = list(kms.values_list('status', flat=True))
                    st_dp = list(despesas.values_list('status', flat=True))
                    todos_status = set(st_km + st_dp)

                    if 'Rejeitado' in todos_status: status_geral = 'Rejeitado'
                    elif 'Pendente' in todos_status: status_geral = 'Pendente'
                    elif 'Aprovado_Regional' in todos_status: status_geral = 'Aprovado_Regional'
                    elif 'Aprovado_Matriz' in todos_status: status_geral = 'Aprovado_Matriz'
                    elif 'Aprovado_Financeiro' in todos_status: status_geral = 'Aprovado_Financeiro'
                    elif 'Pago' in todos_status: status_geral = 'Pago'
                    elif 'Aprovado' in todos_status: status_geral = 'Aprovado_Matriz'
                    
                    ids_km = list(kms.values_list('id', flat=True))

                dados_km_semana_atual.append({
                    'funcionario': f,
                    'total_km': total_km_val,
                    'valor_total_financeiro': valor_total_financeiro,
                    'status': status_geral,
                    'ids_km': ids_km,
                    'tem_registro': tem_algo
                })

    return render(request, 'core_rh/area_gestor.html', {
        'mes_atual': mes,
        'ano_atual': ano,
        'nome_mes': f"{MESES_PT.get(mes, 'Mês')}/{ano}",
        'nav_anterior': nav_anterior, 
        'nav_proximo': nav_proximo,
        'equipes_ponto': equipes_ponto,
        'lista_ponto': lista_ponto,
        'equipes_km': equipes_km,
        'equipe_km_selecionada': equipe_km_selecionada,
        'semanas_do_mes': semanas_do_mes,
        'dados_km_semana_atual': dados_km_semana_atual,
        'semana_selecionada': semana_selecionada,
        'is_gestao': usuario_eh_gestao(request.user),
        'is_financeiro': usuario_eh_financeiro(request.user),        
    })

@login_required
def assinar_ponto_gestor(request, func_id, mes, ano):
    if request.method != 'POST':
        return redirect('area_gestor')
    
    gestor = Funcionario.objects.get(usuario=request.user)
    
    # --- CORREÇÃO: Busca Unificada ---
    equipes_gestor = Equipe.objects.filter(
        Q(oculta=False) & (Q(gestor=gestor) | Q(gestores=gestor))
    )
    
    alvo = Funcionario.objects.get(id=func_id)
    
    # Equipes do alvo (Principal + Secundárias)
    equipes_alvo = [alvo.equipe] + list(alvo.outras_equipes.all())
    
    # Verifica intersecção: Existe alguma equipe que o alvo faz parte E que o gestor lidera?
    is_authorized = any(e in equipes_gestor for e in equipes_alvo if e)

    if not is_authorized and not request.user.is_superuser:
        messages.error(request, "Permissão negada.")
        return redirect('area_gestor')
    
    data_inicio, data_fim = get_datas_competencia(int(mes), int(ano))
    
    if request.FILES.get('arquivo_gestor'):
        arquivo = request.FILES['arquivo_gestor']
        nome_limpo = alvo.nome_completo.strip().replace(' ', '_')
        arquivo.name = f"Folha_{nome_limpo}_{mes}_{ano}_Assinada_Gestor.pdf"
        
        registros = RegistroPonto.objects.filter(
            funcionario=alvo,
            data__range=[data_inicio, data_fim]
        )
        
        if registros.exists():
            
            registros.update(assinado_gestor=True)
            
            registro_com_arquivo = registros.exclude(arquivo_anexo='').first()
            target_reg = registro_com_arquivo if registro_com_arquivo else registros.first()
            
            target_reg.arquivo_anexo = arquivo
            target_reg.save()
            
            messages.success(request, f"Ponto de {alvo.nome_completo} assinado e arquivo atualizado com sucesso!")
        else:
            messages.error(request, "Registros não encontrados para o período.")
    else:
        messages.error(request, "Nenhum arquivo enviado.")
    
    return redirect('area_gestor')

@login_required
def historico_funcionario_view(request, func_id):
    funcionario = Funcionario.objects.get(id=func_id)
    
    historico = []
    hoje = timezone.now()
    
    for i in range(12):
        data_ref = hoje - timedelta(days=i*30)
        mes = data_ref.month
        ano = data_ref.year
        
        m_comp, a_comp = get_competencia_atual() if i == 0 else (mes, ano)
        
        data_ini, data_fim = get_datas_competencia(m_comp, a_comp)
        
        tem_ponto = RegistroPonto.objects.filter(
            funcionario=funcionario,
            data__range=[data_ini, data_fim]
        ).exists()
        
        if tem_ponto:
            historico.append({
                'mes': m_comp,
                'ano': a_comp,
                'nome_mes': f"{MESES_PT[m_comp]}/{a_comp}"
            })

    return render(request, 'core_rh/historico_funcionario.html', {
        'funcionario': funcionario,
        'historico': historico
    })

@login_required
def rh_summary_view(request):
    if not usuario_eh_rh(request.user):
        return HttpResponse("Acesso negado.", status=403)
    
    mes_real, ano_real = get_competencia_atual()
    try:
        mes_solicitado = int(request.GET.get('mes', mes_real))
        ano_solicitado = int(request.GET.get('ano', ano_real))
    except ValueError:
        mes_solicitado = mes_real
        ano_solicitado = ano_real

    mes_ant, ano_ant = get_competencia_anterior(mes_real, ano_real)
    nav_anterior = None
    nav_proximo = None

    if mes_solicitado == mes_real and ano_solicitado == ano_real:
        nav_anterior = {'mes': mes_ant, 'ano': ano_ant}
    elif mes_solicitado == mes_ant and ano_solicitado == ano_ant:
        nav_proximo = {'mes': mes_real, 'ano': ano_real}
    else:
        return redirect(f"{reverse('rh_summary')}?mes={mes_real}&ano={ano_real}")

    data_inicio, data_fim = get_datas_competencia(mes_solicitado, ano_solicitado)
    
    # --- CORREÇÃO: RH vê todas as equipes, MENOS AS OCULTAS ---
    todas_equipes = Equipe.objects.filter(oculta=False).order_by('nome')
    
    resumo_rh = []
    
    for equipe in todas_equipes:
        membros = Funcionario.objects.filter(
            Q(equipe=equipe) | Q(outras_equipes=equipe)
        ).distinct()
        
        total_funcionarios_ativos = membros.count()
        
        membros_com_ponto = RegistroPonto.objects.filter(
            funcionario__in=membros,
            data__range=[data_inicio, data_fim]
        ).values('funcionario').distinct().count()
        
        assinados_gestor = RegistroPonto.objects.filter(
            funcionario__in=membros,
            data__range=[data_inicio, data_fim],
            assinado_gestor=True
        ).values('funcionario').distinct().count()

        resumo_rh.append({
            'equipe': equipe,
            'total_membros': total_funcionarios_ativos,
            'total_pontos_enviados': membros_com_ponto,
            'total_assinados_gestor': assinados_gestor,
            'status_formatado': f"{assinados_gestor}/{membros_com_ponto}" if membros_com_ponto > 0 else "0/0",
            'mes': mes_solicitado,
            'ano': ano_solicitado
        })

    return render(request, 'core_rh/rh_summary.html', {
        'resumo_rh': resumo_rh,
        'mes_atual': MESES_PT.get(mes_solicitado), 
        'mes_num': mes_solicitado, 
        'ano_atual': ano_solicitado,
        'nav_anterior': nav_anterior,
        'nav_proximo': nav_proximo,
    })

@login_required
def rh_team_detail_view(request, equipe_id):
    if not usuario_eh_rh(request.user):
        return HttpResponse("Acesso negado.", status=403)

    equipe = get_object_or_404(Equipe, id=equipe_id)
    
    mes_real, ano_real = get_competencia_atual()
    try:
        mes_solicitado = int(request.GET.get('mes', mes_real))
        ano_solicitado = int(request.GET.get('ano', ano_real))
    except ValueError:
        mes_solicitado = mes_real
        ano_solicitado = ano_real

    mes_ant, ano_ant = get_competencia_anterior(mes_real, ano_real)
    nav_anterior = None
    nav_proximo = None

    if mes_solicitado == mes_real and ano_solicitado == ano_real:
        nav_anterior = {'mes': mes_ant, 'ano': ano_ant}
    elif mes_solicitado == mes_ant and ano_solicitado == ano_ant:
        nav_proximo = {'mes': mes_real, 'ano': ano_real}
    else:
        return redirect(f"{reverse('rh_team_detail', args=[equipe_id])}?mes={mes_real}&ano={ano_real}")

    data_inicio, data_fim = get_datas_competencia(mes_solicitado, ano_solicitado)

    membros = Funcionario.objects.filter(equipe=equipe).order_by('nome_completo')
    lista_colaboradores = []

    for func in membros:
        pontos_mes = RegistroPonto.objects.filter(
            funcionario=func, 
            data__range=[data_inicio, data_fim]
        )
        
        registro_status = pontos_mes.first()
        status_func = registro_status.assinado_funcionario if registro_status else False
        status_gestor = registro_status.assinado_gestor if registro_status else False
        
        anexo_reg = pontos_mes.exclude(arquivo_anexo='').first()
        url_arquivo = anexo_reg.arquivo_anexo.url if anexo_reg and anexo_reg.arquivo_anexo else None
        
        nome_limpo = func.nome_completo.strip().replace(' ', '_')
        nome_para_download = f"Folha_{nome_limpo}_{mes_solicitado:02d}_{ano_solicitado}.pdf"

        lista_colaboradores.append({
            'funcionario': func,
            'status_func': status_func,
            'status_gestor': status_gestor,
            'arquivo_anexo': url_arquivo,
            'nome_download': nome_para_download,
            'mes': mes_solicitado,
            'ano': ano_solicitado,
        })

    return render(request, 'core_rh/rh_team_detail.html', {
        'equipe': equipe,
        'lista_colaboradores': lista_colaboradores,
        'mes_atual': MESES_PT.get(mes_solicitado),
        'mes_num': mes_solicitado,
        'ano_atual': ano_solicitado,
        'nav_anterior': nav_anterior,
        'nav_proximo': nav_proximo,
    })

def rh_batch_download_view(request, equipe_id):
    """
    Gera um ZIP com todos os PDFs assinados da equipe no mês selecionado.
    Em caso de erro, redireciona de volta para a página atual com um alerta.
    """
    equipe = get_object_or_404(Equipe, pk=equipe_id)
    mes = request.GET.get('mes')
    ano = request.GET.get('ano')

    if not mes or not ano:
        messages.error(request, "Mês e Ano não informados para download.")
        return redirect(request.META.get('HTTP_REFERER', '/'))

    registros = RegistroPonto.objects.filter(
        funcionario__equipe=equipe,
        data__month=mes,
        data__year=ano
    ).exclude(arquivo_anexo='').exclude(arquivo_anexo__isnull=True)

    if not registros.exists():
        messages.warning(request, f"Nenhum ponto assinado encontrado para a equipe {equipe.nome} em {mes}/{ano}.")
        return redirect(request.META.get('HTTP_REFERER', '/'))

    zip_buffer = io.BytesIO()
    arquivos_adicionados = 0

    with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
        for ponto in registros:
            try:
                file_path = ponto.arquivo_anexo.path
                if os.path.exists(file_path):
                    file_name = f"{ponto.funcionario.nome_completo}_{ponto.data.strftime('%d-%m-%Y')}.pdf"
                    zip_file.write(file_path, file_name)
                    arquivos_adicionados += 1
            except Exception as e:
                print(f"Erro ao adicionar arquivo {ponto}: {e}")
                continue

    if arquivos_adicionados == 0:
        messages.error(request, "Arquivos físicos não encontrados no servidor.")
        return redirect(request.META.get('HTTP_REFERER', '/'))

    zip_buffer.seek(0)
    response = HttpResponse(zip_buffer, content_type='application/zip')
    nome_zip = f"Pontos_{equipe.nome}_{mes}_{ano}.zip"
    response['Content-Disposition'] = f'attachment; filename="{nome_zip}"'
    
    return response

@login_required
def rh_unlock_timesheet_view(request, func_id, mes, ano):
    if not usuario_eh_rh(request.user):
        return HttpResponse("Acesso negado. Perfil RH necessário.", status=403)
        
    funcionario = get_object_or_404(Funcionario, id=func_id)
    data_inicio, data_fim = get_datas_competencia(mes, ano)

    registros = RegistroPonto.objects.filter(
        funcionario=funcionario,
        data__range=[data_inicio, data_fim]
    )

    if registros.exists():
        registros.update(assinado_gestor=False)
        messages.success(request, f"Folha de {funcionario.nome_completo} desbloqueada com sucesso!")
    else:
        messages.error(request, "Nenhum registro encontrado para desbloquear.")

    return redirect(request.META.get('HTTP_REFERER', '/'))

@login_required
def trocar_senha_obrigatoria(request):
    try:
        funcionario = request.user.funcionario
        if not funcionario.primeiro_acesso:
            return redirect('home')
    except Funcionario.DoesNotExist:
        return redirect('home')

    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            funcionario.primeiro_acesso = False
            funcionario.save()
            messages.success(request, 'Senha atualizada com sucesso! Bem-vindo.')
            return redirect('home')
        else:
            messages.error(request, 'Por favor, corrija os erros abaixo.')
    else:
        form = PasswordChangeForm(request.user)
        
    return render(request, 'core_rh/trocar_senha.html', {'form': form})

def format_delta(td):
    total_seconds = int(td.total_seconds())
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    return f"{hours:02}:{minutes:02}"

@login_required
def admin_ponto_partial_view(request, func_id):
    if not usuario_eh_rh(request.user):
        return HttpResponse("Acesso negado", status=403)

    funcionario = get_object_or_404(Funcionario, pk=func_id)
    
    mes_real, ano_real = get_competencia_atual()
    try:
        mes = int(request.GET.get('mes', mes_real))
        ano = int(request.GET.get('ano', ano_real))
    except ValueError:
        mes, ano = mes_real, ano_real

    mes_ant, ano_ant = get_competencia_anterior(mes, ano)
    
    if mes == 12:
        mes_prox, ano_prox = 1, ano + 1
    else:
        mes_prox, ano_prox = mes + 1, ano

    data_inicio, data_fim = get_datas_competencia(mes, ano)
    
    registros = RegistroPonto.objects.filter(
        funcionario=funcionario,
        data__range=[data_inicio, data_fim]
    ).order_by('data')

    registros_dict = {r.data.day: r for r in registros}
    dias_do_mes = []
    
    feriados_br = holidays.BR(state='DF', years=ano)
    if hasattr(funcionario, 'estado_sigla') and funcionario.estado_sigla:
         try:
             feriados_br = holidays.BR(state=funcionario.estado_sigla, years=ano)
         except: pass

    delta_dias = (data_fim - data_inicio).days
    for i in range(delta_dias + 1):
        data_atual = data_inicio + timedelta(days=i)
        eh_feriado = data_atual in feriados_br
        
        dias_do_mes.append({
            'data': data_atual,
            'dia_semana_nome': DIAS_SEMANA_PT[data_atual.weekday()],
            'eh_feriado': eh_feriado,
            'nome_feriado': feriados_br.get(data_atual).upper() if eh_feriado else "",
            'registro': registros_dict.get(data_atual.day)
        })

    context = {
        'funcionario': funcionario,
        'dias_do_mes': dias_do_mes,
        'mes_atual': mes,
        'ano_atual': ano,
        'nome_mes': f"{MESES_PT.get(mes)}/{ano}",
        'mes_anterior': mes_ant,
        'ano_anterior': ano_ant,
        'mes_proximo': mes_prox,
        'ano_proximo': ano_prox,
        'is_admin_view': True, 
    }

    return render(request, 'core_rh/includes/folha_ponto_content.html', context)


@login_required
def admin_gestor_partial_view(request):
    # 1. Identificação e Permissão Unificada
    user = request.user
    is_rh = usuario_eh_rh(user)
    funcionario = None
    
    try:
        funcionario = user.funcionario
    except AttributeError:
        if not is_rh: return HttpResponse("Negado", status=403)

    # Definição das Equipes Permitidas
    if is_rh:
        equipes_permitidas = Equipe.objects.filter(oculta=False).order_by('nome')
    else:
        equipes_permitidas = Equipe.objects.filter(
            Q(oculta=False) & (Q(gestor=funcionario) | Q(gestores=funcionario))
        ).distinct().order_by('nome')

    if not is_rh and not equipes_permitidas.exists():
        return HttpResponse('<div class="alert alert-warning">Você não gerencia nenhuma equipe ativa.</div>', status=403)

    # 2. Filtros e Datas (Mensais - Padrão Ponto)
    ma, aa = get_competencia_atual()
    try: 
        m = int(request.GET.get('mes', ma))
        a = int(request.GET.get('ano', aa))
        mode = request.GET.get('mode', 'list')
        eq_id = request.GET.get('equipe_id', '')
        q = request.GET.get('q', '').strip()
        est = request.GET.get('estado', '')
    except: m, a, mode, eq_id, q, est = ma, aa, 'list', '', '', ''
    
    # Navegação Mês
    mp, ap = get_competencia_anterior(m, a)
    mpr, apr = (1, a+1) if m==12 else (m+1, a)
    nav_ant = {'mes': mp, 'ano': ap}
    nav_prox = {'mes': mpr, 'ano': apr}
    
    # Datas do Mês (Para Ponto)
    di_mes, df_mes = get_datas_competencia(m, a)
    
    ctx = {
        'mes_atual': m, 'ano_atual': a, 
        'nome_mes': f"{MESES_PT.get(m)}/{a}", 
        'nav_anterior': nav_ant, 'nav_proximo': nav_prox, 
        'mode': mode, 'q': q, 'equipe_id': eq_id, 'estado_filtro': est,
        'todas_equipes': equipes_permitidas,
        'is_gestao': usuario_eh_rh(user), # Flag para botões de aprovação
        'is_financeiro': user.groups.filter(name='Financeiro').exists() or user.is_superuser
    }
    
    # =========================================================================
    # LÓGICA 1: GESTÃO DE PONTO (ABA 1)
    # =========================================================================
    if mode == 'summary':
        # Modo Cards (Resumo)
        qs_resumo = equipes_permitidas
        if q: qs_resumo = qs_resumo.filter(nome__icontains=q)
        
        res = []
        for e in qs_resumo:
            mem = Funcionario.objects.filter(Q(equipe=e)|Q(outras_equipes=e)).distinct()
            ass = RegistroPonto.objects.filter(funcionario__in=mem, data__range=[di_mes, df_mes], assinado_gestor=True).values('funcionario').distinct().count()
            tot = mem.count()
            res.append({'equipe': e, 'total_membros': tot, 'total_assinados': ass, 'progresso': int(ass/tot*100) if tot>0 else 0})
        ctx['resumo_rh'] = res
        
    else:
        # Modo Lista (Tabela de Funcionários do Ponto)
        fq = Funcionario.objects.filter(usuario__is_active=True)
        if est: fq = fq.filter(local_trabalho_estado=est)
        
        # Filtro de Equipe (Ponto)
        if eq_id: 
            if equipes_permitidas.filter(id=eq_id).exists():
                fq = fq.filter(Q(equipe_id=eq_id)|Q(outras_equipes__id=eq_id))
            else:
                fq = fq.none()
        else:
            fq = fq.filter(Q(equipe__in=equipes_permitidas)|Q(outras_equipes__in=equipes_permitidas))
            
        if q: fq = fq.filter(nome_completo__icontains=q)
        
        lst = []
        for f in fq.distinct().order_by('nome_completo'):
            # Verifica status do ponto no mês
            pts = RegistroPonto.objects.filter(funcionario=f, data__range=[di_mes, df_mes])
            tem_assinatura_func = pts.filter(assinado_funcionario=True).exists()
            tem_assinatura_gest = pts.filter(assinado_gestor=True).exists()
            
            # Pega link do arquivo se existir
            arq_url = None
            ponto_final = pts.exclude(arquivo_anexo='').order_by('-data').first()
            if ponto_final and ponto_final.arquivo_anexo:
                arq_url = ponto_final.arquivo_anexo.url

            lst.append({
                'funcionario': f, 
                'status_func': tem_assinatura_func, 
                'status_gestor': tem_assinatura_gest, 
                'pode_assinar': (tem_assinatura_func and not tem_assinatura_gest),
                'arquivo_assinado_url': arq_url, 
                'nome_download': f"Folha_{f.nome_completo.split()[0]}_{m}_{a}.pdf", 
                'mes': m, 'ano': a
            })
        ctx['lista_colaboradores'] = lst # Usado na aba Ponto
        ctx['lista_ponto'] = lst # Alias para garantir compatibilidade
        ctx['estados_disponiveis'] = fq.exclude(local_trabalho_estado__isnull=True).values_list('local_trabalho_estado', flat=True).distinct()

    # =========================================================================
    # LÓGICA 2: GESTÃO DE KM / CAMPO (ABA 2)
    # =========================================================================
    
    # 1. Filtrar apenas equipes de "Campo" (Contém 'Campo' no nome)
    equipes_km = equipes_permitidas.filter(nome__icontains="Campo")
    ctx['equipes_km'] = equipes_km

    if equipes_km.exists():
        # 2. Definir Equipe Selecionada para KM
        km_team_id = request.GET.get('km_team')
        if km_team_id and equipes_km.filter(id=km_team_id).exists():
            equipe_km_selecionada = equipes_km.get(id=km_team_id)
        else:
            equipe_km_selecionada = equipes_km.first()
        
        ctx['equipe_km_selecionada'] = equipe_km_selecionada

        # 3. Calcular Semanas do Mês
        cal = monthcalendar(a, m)
        semanas_validas = [s for s in cal if any(d != 0 for d in s)]
        
        try:
            semana_param = int(request.GET.get('semana', 1))
        except:
            semana_param = 1
            
        # Garante índice válido
        if semana_param < 1: semana_param = 1
        if semana_param > len(semanas_validas): semana_param = len(semanas_validas)

        # Monta lista de semanas para as pílulas de navegação
        semanas_info = []
        dt_inicio_sem = None
        dt_fim_sem = None

        for idx, semana_lista in enumerate(semanas_validas, start=1):
            dia_ref = next(d for d in semana_lista if d != 0)
            data_ref = date(a, m, dia_ref)
            # Segunda a Domingo
            inicio = data_ref - timedelta(days=data_ref.weekday())
            fim = inicio + timedelta(days=6)
            
            active = (idx == semana_param)
            if active:
                dt_inicio_sem = inicio
                dt_fim_sem = fim
            
            semanas_info.append({
                'numero': idx,
                'inicio': inicio,
                'fim': fim,
                'active': active
            })
        
        ctx['semanas_do_mes'] = semanas_info
        ctx['semana_selecionada'] = semana_param

        # 4. Buscar Dados Financeiros (KM + Despesas) da Semana Selecionada
        funcionarios_km = Funcionario.objects.filter(Q(equipe=equipe_km_selecionada)|Q(outras_equipes=equipe_km_selecionada)).distinct().order_by('nome_completo')
        
        dados_km_semana = []
        
        for func in funcionarios_km:
            # Filtra registros da semana específica
            kms = ControleKM.objects.filter(funcionario=func, data__range=[dt_inicio_sem, dt_fim_sem])
            despesas = DespesaDiversa.objects.filter(funcionario=func, data__range=[dt_inicio_sem, dt_fim_sem])
            
            tem_registro = kms.exists() or despesas.exists()
            
            if tem_registro:
                # Cálculos
                total_km_val = sum(k.total_km for k in kms)
                total_despesas_val = sum(d.valor for d in despesas)
                
                # Fator multiplicador (Se não tiver no func, usa 1.20)
                fator = float(func.valor_km) if func.valor_km and func.valor_km > 0 else 1.20
                
                # CÁLCULO FINAL: (KM * Fator) + Despesas
                valor_total_financeiro = (float(total_km_val) * fator) + float(total_despesas_val)
                
                # Definição de Status Geral (Prioridade: Rejeitado > Pendente > Aprovado)
                todos_status = list(kms.values_list('status', flat=True)) + list(despesas.values_list('status', flat=True))
                
                if 'Rejeitado' in todos_status: status_geral = 'Rejeitado'
                elif 'Pendente' in todos_status: status_geral = 'Pendente'
                elif 'Aprovado_Regional' in todos_status: status_geral = 'Aprovado_Regional'
                elif 'Aprovado_Matriz' in todos_status: status_geral = 'Aprovado_Matriz'
                elif 'Aprovado_Financeiro' in todos_status: status_geral = 'Aprovado_Financeiro'
                elif 'Pago' in todos_status: status_geral = 'Pago'
                else: status_geral = 'Vazio'

                dados_km_semana.append({
                    'funcionario': func,
                    'tem_registro': True,
                    'total_km': total_km_val,
                    'valor_total_financeiro': valor_total_financeiro, # CAMPO CALCULADO
                    'status': status_geral,
                    'ids_km': list(kms.values_list('id', flat=True)) # Para links de ação
                })
            else:
                # Opcional: Mostrar funcionários sem registro na lista?
                # Se não quiser mostrar quem não rodou, comente o append abaixo
                # dados_km_semana.append({'funcionario': func, 'tem_registro': False, 'status': 'Vazio'})
                pass

        ctx['dados_km_semana_atual'] = dados_km_semana

    return render(request, 'core_rh/includes/rh_area_moderno.html', ctx)
try:
    from .models import Ferias
except ImportError:
    pass

@login_required
def minhas_ferias_view(request):
    try:
        funcionario = Funcionario.objects.get(usuario=request.user)
    except Funcionario.DoesNotExist:
        return redirect('home')
    
    lista_ferias = Ferias.objects.filter(funcionario=funcionario).order_by('-data_inicio')
    
    return render(request, 'core_rh/ferias.html', {
        'funcionario': funcionario,
        'lista_ferias': lista_ferias
    })

@login_required
def upload_ferias_view(request, ferias_id):
    if request.method != 'POST': return redirect('minhas_ferias')
    
    ferias = get_object_or_404(Ferias, id=ferias_id, funcionario__usuario=request.user)
    updated = False
    
    if request.FILES.get('aviso_file'):
        ferias.aviso_assinado = request.FILES['aviso_file']
        updated = True
        
    if request.FILES.get('recibo_file'):
        ferias.recibo_assinado = request.FILES['recibo_file']
        updated = True
        
    if updated:
        if ferias.status == 'Pendente': ferias.status = 'Enviado'
        ferias.save()
        messages.success(request, "Arquivo enviado com sucesso!")
        
    return redirect('minhas_ferias')

@login_required
def gerar_aviso_ferias_pdf(request, ferias_id):
    if not (request.user.is_staff or usuario_eh_rh(request.user)):
        return redirect('home')
        
    ferias = get_object_or_404(Ferias, id=ferias_id)
    func = ferias.funcionario
    
    dias_ferias = (ferias.data_fim - ferias.data_inicio).days + 1
    
    html_string = render_to_string('core_rh/pdf_aviso_ferias.html', {
        'ferias': ferias,
        'func': func,
        'dias_ferias': dias_ferias,
        'hoje': timezone.now()
    })

    try:
        from weasyprint import HTML
        html = HTML(string=html_string)
        pdf_file = html.write_pdf(optimize_size=('fonts', 'images'))
        
        nome_func = func.nome_completo.strip().replace(' ', '_')
        periodo_limpo = ferias.periodo_aquisitivo.replace('/', '-')
        filename = f"Notificação_de_Férias-{nome_func}-{periodo_limpo}.pdf"

        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    except ImportError:
        return HttpResponse(html_string)

@login_required
def admin_ferias_partial_view(request):
    if not (request.user.is_staff or usuario_eh_rh(request.user)):
        return HttpResponse("Acesso negado", status=403)
    
    q = request.GET.get('q', '').strip()
    status_filtro = request.GET.get('status', '')
    mes = request.GET.get('mes')
    ano = request.GET.get('ano')

    hoje = timezone.now().date()
    try:
        ano = int(ano) if ano else hoje.year
        mes = int(mes) if mes else hoje.month
        data_base = date(ano, mes, 1)
    except:
        data_base = hoje

    mes_ant = data_base.month - 1 if data_base.month > 1 else 12
    ano_ant = data_base.year if data_base.month > 1 else data_base.year - 1
    
    mes_prox = data_base.month + 1 if data_base.month < 12 else 1
    ano_prox = data_base.year if data_base.month < 12 else data_base.year + 1

    ultimo_dia_mes = monthrange(data_base.year, data_base.month)[1]
    data_fim_mes = date(data_base.year, data_base.month, ultimo_dia_mes)
    
    from .models import Ferias
    
    ferias_qs = Ferias.objects.filter(
        data_inicio__lte=data_fim_mes,
        data_fim__gte=data_base
    ).select_related('funcionario', 'funcionario__equipe')

    if q:
        ferias_qs = ferias_qs.filter(
            Q(funcionario__nome_completo__icontains=q) | 
            Q(funcionario__matricula__icontains=q)
        )

    if status_filtro:
        ferias_qs = ferias_qs.filter(status=status_filtro)

    ferias_qs = ferias_qs.order_by('data_inicio')

    context = {
        'lista_ferias': ferias_qs,
        'mes_atual': data_base.month,
        'ano_atual': data_base.year,
        'nav_anterior': {'mes': mes_ant, 'ano': ano_ant},
        'nav_proximo': {'mes': mes_prox, 'ano': ano_prox},
        'q': q,
        'status_filtro': status_filtro,
    }
    
    meses = {1:'Janeiro', 2:'Fevereiro', 3:'Março', 4:'Abril', 5:'Maio', 6:'Junho', 
             7:'Julho', 8:'Agosto', 9:'Setembro', 10:'Outubro', 11:'Novembro', 12:'Dezembro'}
    context['nome_mes'] = f"{meses.get(data_base.month)} {data_base.year}"

    return render(request, 'core_rh/includes/rh_ferias_moderno.html', context)

@login_required
def meus_contracheques(request):
    try:
        funcionario = request.user.funcionario
        lista = Contracheque.objects.filter(funcionario=funcionario).order_by('-ano', '-mes')
    except AttributeError:
        lista = []
        messages.error(request, "Seu usuário não está vinculado a um funcionário.")

    return render(request, 'core_rh/meus_contracheques.html', {'lista': lista})

@login_required
def assinar_contracheque_local(request, pk):
    if request.method == "POST":
        contracheque = get_object_or_404(Contracheque, pk=pk, funcionario__usuario=request.user)
        
        if not contracheque.data_ciencia:
            try:
                funcionario = contracheque.funcionario
                nome_assinatura = funcionario.nome_completo.strip().upper()
                
                pdf_io = io.BytesIO(contracheque.arquivo.read())
                
                plumber_pdf = pdfplumber.open(pdf_io)
                pdf_io.seek(0)
                reader = PdfReader(pdf_io)
                writer = PdfWriter()
                
                for i, page in enumerate(reader.pages):
                    try:
                        p_page = plumber_pdf.pages[i]
                        altura_pagina = float(page.mediabox.height)
                        
                        palavras = p_page.search("ASSINATURA") or \
                                   p_page.search("EMPREGADO")
                        
                        pos_x = 400
                        pos_y_base = 50
                        largura_final = 250
                        
                        if palavras:
                            target = palavras[-1] 
                            centro_texto = (target['x0'] + target['x1']) / 2
                            
                            linha_exata = None
                            
                            limite_busca_inferior = target['top'] - 25 
                            limite_busca_superior = target['top'] 
                            
                            for linha in p_page.lines:
                                if abs(linha['top'] - linha['bottom']) < 2:
                                    
                                    if limite_busca_inferior < linha['bottom'] < limite_busca_superior:
                                        
                                        if linha['x0'] < centro_texto < linha['x1']:
                                            linha_exata = linha
                                            break 
                            
                            if linha_exata:
                                pos_x = linha_exata['x0']
                                largura_final = linha_exata['x1'] - linha_exata['x0']
                                
                                pos_y_base = altura_pagina - linha_exata['bottom']
                            else:
                                largura_texto = target['x1'] - target['x0']
                                largura_final = largura_texto * 3 
                                if largura_final < 200: largura_final = 200 
                                
                                pos_x = centro_texto - (largura_final / 2)
                                pos_y_base = altura_pagina - target['bottom']

                        packet = io.BytesIO()
                        can = canvas.Canvas(packet, pagesize=A4)
                        
                        rect_h = 32 
                        rect_y = pos_y_base - 2 
                        
                        can.setFillColor(colors.white)
                        can.rect(pos_x, rect_y, largura_final, rect_h, stroke=0, fill=1)
                        
                        font_size = 10
                        nome_width = can.stringWidth(nome_assinatura, "Helvetica-Bold", font_size)
                        
                        if nome_width > largura_final:
                            font_size = font_size * (largura_final / nome_width) * 0.95
                        
                        can.setFillColor(colors.black)
                        can.setFont("Helvetica-Bold", font_size)
                        
                        centro_area = pos_x + (largura_final / 2)
                        can.drawCentredString(centro_area, rect_y + 12, nome_assinatura)
                        
                        can.setLineWidth(0.5)
                        y_linha = rect_y + 10
                        can.line(pos_x, y_linha, pos_x + largura_final, y_linha)
                        
                        can.setFont("Helvetica", 6)
                        can.drawCentredString(centro_area, rect_y + 2, "ASSINATURA")
                        
                        can.save()
                        packet.seek(0)
                        overlay = PdfReader(packet)
                        page.merge_page(overlay.pages[0])
                        
                    except Exception as e:
                        print(f"Erro processando página {i}: {e}")
                    
                    writer.add_page(page)
                
                plumber_pdf.close()
                
                pdf_output = io.BytesIO()
                writer.write(pdf_output)
                
                filename = f"holerite_{funcionario.id}_{contracheque.mes}_{contracheque.ano}_assinado.pdf"
                contracheque.arquivo.save(filename, ContentFile(pdf_output.getvalue()), save=False)
                
                contracheque.data_ciencia = timezone.now()
                ip = request.META.get('HTTP_X_FORWARDED_FOR', request.META.get('REMOTE_ADDR')).split(',')[0]
                contracheque.ip_ciencia = ip
                contracheque.save()
                
                messages.success(request, "Assinado com sucesso!")
                
            except Exception as e:
                messages.error(request, f"Erro: {str(e)}")
        
        return redirect('meus_contracheques')

@login_required
def admin_contracheque_partial(request):
    hoje = timezone.now()
    try:
        mes_atual = int(request.GET.get('mes', hoje.month))
        ano_atual = int(request.GET.get('ano', hoje.year))
    except ValueError:
        mes_atual = hoje.month
        ano_atual = hoje.year

    termo_busca = request.GET.get('q', '').strip()

    mes_anterior = mes_atual - 1 if mes_atual > 1 else 12
    ano_anterior = ano_atual if mes_atual > 1 else ano_atual - 1
    
    mes_proximo = mes_atual + 1 if mes_atual < 12 else 1
    ano_proximo = ano_atual if mes_atual < 12 else ano_atual + 1

    funcionarios = Funcionario.objects.all().order_by('nome_completo')
    
    if termo_busca:
        funcionarios = funcionarios.filter(nome_completo__icontains=termo_busca)

    lista_equipe = []
    
    contracheques_mes = {
        cc.funcionario_id: cc 
        for cc in Contracheque.objects.filter(mes=mes_atual, ano=ano_atual)
    }

    for func in funcionarios:
        cc = contracheques_mes.get(func.id)
        
        status_envio = bool(cc and cc.arquivo)
        status_assinatura = bool(cc and cc.data_ciencia)
        
        lista_equipe.append({
            'funcionario': func,
            'contracheque': cc,
            'enviado': status_envio,
            'assinado': status_assinatura
        })

    context = {
        'lista_equipe': lista_equipe,
        'mes_atual': mes_atual,
        'ano_atual': ano_atual,
        'nome_mes': dict(Contracheque.MESES).get(mes_atual),
        'nav_anterior': {'mes': mes_anterior, 'ano': ano_anterior},
        'nav_proximo': {'mes': mes_proximo, 'ano': ano_proximo},
        'q': termo_busca,
        'meses_choices': Contracheque.MESES, 
    }
    
    return render(request, 'core_rh/includes/rh_contracheque_moderno.html', context)

@login_required
def gerenciar_contracheques(request):
    try:
        fallback_url = reverse('admin:core_rh_funcionario_changelist')
    except:
        fallback_url = '/admin/core_rh/funcionario/'
        
    next_url = request.POST.get('next') or request.GET.get('next') or fallback_url

    if not (request.user.is_staff or usuario_eh_rh(request.user)):
        return render(request, 'core_rh/upload_log.html', {'erro_critico': 'Acesso Negado.', 'next_url': next_url})

    if request.method == 'GET':
        return redirect(next_url)

    if 'arquivo_pdf' not in request.FILES:
        return render(request, 'core_rh/upload_log.html', {'erro_critico': 'Nenhum arquivo enviado.', 'next_url': next_url})

    try:
        arquivo = request.FILES['arquivo_pdf']
        mes_upload = int(request.POST.get('mes_upload'))
        ano_upload = int(request.POST.get('ano_upload'))
        
        data_str = request.POST.get('data_recebimento')
        data_para_pdf = None
        if data_str:
            data_para_pdf = datetime.strptime(data_str, '%Y-%m-%d').date()

        arquivo.seek(0)
        plumber_pdf = pdfplumber.open(arquivo)
        arquivo.seek(0)
        reader = PdfReader(arquivo)

        funcionarios_db = Funcionario.objects.all()
        log_sucesso = []
        log_erro = []
        
        for i, page in enumerate(reader.pages):
            texto = page.extract_text() or ""
            texto_upper = texto.upper()
            
            encontrou_func = False
            funcionario_encontrado = None
            for func in funcionarios_db:
                nome_busca = func.nome_completo.strip().upper()
                if nome_busca and nome_busca in texto_upper:
                    funcionario_encontrado = func
                    encontrou_func = True
                    break 
            
            if encontrou_func:
                writer = PdfWriter()
                
                if data_para_pdf:
                    try:
                        p_page = plumber_pdf.pages[i]
                        altura_pagina = float(page.mediabox.height)

                        palavras_alvo = p_page.search("DATA DO RECEBIMENTO") or \
                                        p_page.search("DATA RECEBIMENTO") or \
                                        p_page.search("RECEBIMENTO")
                        
                        palavras_limite = p_page.search("Declaro ter recebido") or \
                                          p_page.search("Declaro")

                        if palavras_alvo:
                            target = palavras_alvo[0]
                            if palavras_limite:
                                limite_top = palavras_limite[0]['bottom']
                            else:
                                limite_top = target['top'] - 40

                            limite_bottom = target['bottom']
                            rect_y = altura_pagina - limite_bottom
                            rect_h = limite_bottom - limite_top
                            rect_x = target['x0'] - 20 
                            rect_w = (target['x1'] - target['x0']) + 40 

                            packet = io.BytesIO()
                            can = canvas.Canvas(packet, pagesize=A4)
                            
                            can.setFillColor(colors.white)
                            can.rect(rect_x, rect_y - 2, rect_w, rect_h + 4, stroke=0, fill=1)
                            
                            altura_legenda = target['bottom'] - target['top']
                            tamanho_fonte_data = altura_legenda * 1.8
                            if tamanho_fonte_data < 12: tamanho_fonte_data = 12
                            if tamanho_fonte_data > 18: tamanho_fonte_data = 18

                            pos_data_x = target['x0']
                            pos_data_y = rect_y + 10 

                            can.setFillColor(colors.black)
                            can.setFont("Helvetica-Bold", tamanho_fonte_data)
                            can.drawString(pos_data_x, pos_data_y, data_para_pdf.strftime("%d/%m/%Y"))

                            can.setLineWidth(0.5)
                            can.line(rect_x, rect_y + 8, rect_x + rect_w, rect_y + 8)
                            can.setFont("Helvetica", 6)
                            can.drawCentredString(rect_x + (rect_w/2), rect_y, "DATA DO RECEBIMENTO")
                            
                            can.save()
                            packet.seek(0)
                            overlay = PdfReader(packet)
                            page.merge_page(overlay.pages[0])
                    except Exception:
                        pass

                writer.add_page(page)
                pdf_bytes = io.BytesIO()
                writer.write(pdf_bytes)
                
                cc, created = Contracheque.objects.update_or_create(
                    funcionario=funcionario_encontrado, mes=mes_upload, ano=ano_upload,
                    defaults={'arquivo': None}
                )
                
                nome_arq = f"holerite_{funcionario_encontrado.id}_{mes_upload}_{ano_upload}.pdf"
                cc.arquivo.save(nome_arq, ContentFile(pdf_bytes.getvalue()))
                
                log_sucesso.append({
                    'pagina': i + 1,
                    'nome': funcionario_encontrado.nome_completo,
                    'status': 'Processado'
                })
            else:
                log_erro.append({'pagina': i + 1, 'motivo': 'Nome não encontrado.'})
        
        plumber_pdf.close()

        return render(request, 'core_rh/upload_log.html', {
            'log_sucesso': log_sucesso,
            'log_erro': log_erro,
            'total_sucesso': len(log_sucesso),
            'total_erro': len(log_erro),
            'mes_nome': dict(Contracheque.MESES).get(mes_upload, mes_upload),
            'ano': ano_upload,
            'next_url': next_url, 
        })

    except Exception as e:
        return render(request, 'core_rh/upload_log.html', {'erro_critico': str(e), 'next_url': next_url})

@login_required
def upload_individual_contracheque(request, func_id):
    if not (request.user.is_staff or usuario_eh_rh(request.user)):
        return HttpResponse("Acesso negado", status=403)
        
    if request.method == 'POST' and request.FILES.get('arquivo_individual'):
        funcionario = get_object_or_404(Funcionario, pk=func_id)
        arquivo = request.FILES['arquivo_individual']
        
        try:
            mes = int(request.POST.get('mes'))
            ano = int(request.POST.get('ano'))
            
            data_str = request.POST.get('data_recebimento_individual')
            data_para_pdf = None
            if data_str:
                data_para_pdf = datetime.strptime(data_str, '%Y-%m-%d').date()
                
        except (ValueError, TypeError):
            messages.error(request, "Dados inválidos.")
            return redirect(request.META.get('HTTP_REFERER', '/'))

        try:
            arquivo.seek(0)
            plumber_pdf = pdfplumber.open(arquivo)
            arquivo.seek(0)
            reader = PdfReader(arquivo)
            writer = PdfWriter()

            for i, page in enumerate(reader.pages):
                if data_para_pdf:
                    try:
                        p_page = plumber_pdf.pages[i]
                        altura_pagina = float(page.mediabox.height)
                        
                        palavras_alvo = p_page.search("DATA DO RECEBIMENTO") or p_page.search("RECEBIMENTO")
                        palavras_limite = p_page.search("Declaro ter recebido") or p_page.search("Declaro")

                        if palavras_alvo:
                            target = palavras_alvo[0]
                            limite_top = palavras_limite[0]['bottom'] if palavras_limite else target['top'] - 40
                            limite_bottom = target['bottom']
                            
                            rect_y = altura_pagina - limite_bottom
                            rect_h = limite_bottom - limite_top
                            rect_x = target['x0'] - 20 
                            rect_w = (target['x1'] - target['x0']) + 40 

                            packet = io.BytesIO()
                            can = canvas.Canvas(packet, pagesize=A4)
                            
                            can.setFillColor(colors.white)
                            can.rect(rect_x, rect_y - 2, rect_w, rect_h + 4, stroke=0, fill=1)
                            
                            altura_legenda = target['bottom'] - target['top']
                            tamanho = max(12, min(18, altura_legenda * 1.8))
                            
                            can.setFillColor(colors.black)
                            can.setFont("Helvetica-Bold", tamanho)
                            can.drawString(target['x0'], rect_y + 10, data_para_pdf.strftime("%d/%m/%Y"))

                            can.setLineWidth(0.5)
                            can.line(rect_x, rect_y + 8, rect_x + rect_w, rect_y + 8)
                            can.setFont("Helvetica", 6)
                            can.drawCentredString(rect_x + (rect_w/2), rect_y, "DATA DO RECEBIMENTO")
                            
                            can.save()
                            packet.seek(0)
                            overlay = PdfReader(packet)
                            page.merge_page(overlay.pages[0])
                    except: pass
                
                writer.add_page(page)
            
            plumber_pdf.close()
            pdf_bytes = io.BytesIO()
            writer.write(pdf_bytes)
            
            cc, created = Contracheque.objects.update_or_create(
                funcionario=funcionario, mes=mes, ano=ano,
                defaults={'arquivo': None}
            )
            
            nome_arq = f"holerite_{funcionario.id}_{mes}_{ano}_manual.pdf"
            cc.arquivo.save(nome_arq, ContentFile(pdf_bytes.getvalue()))
            
            messages.success(request, f"Contracheque de {funcionario.nome_completo} anexado!")

        except Exception as e:
            messages.error(request, f"Erro ao processar arquivo: {e}")
        
    return redirect(request.META.get('HTTP_REFERER', '/'))

@login_required
def excluir_contracheque(request, cc_id):
    if not (request.user.is_staff or usuario_eh_rh(request.user)):
        return HttpResponse("Acesso negado", status=403)
        
    contracheque = get_object_or_404(Contracheque, pk=cc_id)
    nome = contracheque.funcionario.nome_completo
    contracheque.delete()
    
    messages.success(request, f"Contracheque de {nome} removido com sucesso.")
    return redirect(request.META.get('HTTP_REFERER', '/'))

@login_required
def meus_atestados_view(request):
    try:
        funcionario = request.user.funcionario
    except AttributeError:
        return redirect('home')

    if request.method == 'POST':
        try:
            tipo = request.POST.get('tipo')
            data = request.POST.get('data_inicio')
            arquivo = request.FILES.get('arquivo')
            motivo = request.POST.get('motivo')
            
            atestado = Atestado(
                funcionario=funcionario,
                tipo=tipo,
                data_inicio=data,
                motivo=motivo,
                arquivo=arquivo
            )
            
            if tipo == 'DIAS':
                atestado.qtd_dias = int(request.POST.get('qtd_dias'))
            else: 
                atestado.hora_inicio = request.POST.get('hora_inicio')
                atestado.hora_fim = request.POST.get('hora_fim')
                atestado.qtd_dias = 0 
            
            atestado.save()
            messages.success(request, "Documento enviado com sucesso! Aguarde análise do RH.")
            
        except Exception as e:
            messages.error(request, f"Erro ao salvar: {e}")
            
        return redirect('meus_atestados')

    lista = Atestado.objects.filter(funcionario=funcionario).order_by('-data_envio')
    return render(request, 'core_rh/meus_atestados.html', {'lista': lista, 'funcionario': funcionario})

@login_required
def rh_gestao_atestados(request):
    if not usuario_eh_rh(request.user):
        return HttpResponse("Acesso Negado", status=403)
        
    if request.method == 'POST':
        atestado_id = request.POST.get('atestado_id')
        acao = request.POST.get('acao')
        obs = request.POST.get('observacao_rh')
        
        atestado = get_object_or_404(Atestado, id=atestado_id)
        
        if acao == 'aprovar':
            atestado.status = 'Aprovado'
            atestado.observacao_rh = obs
            atestado.save()
            messages.success(request, f"Atestado de {atestado.funcionario.nome_completo} APROVADO.")
        elif acao == 'recusar':
            atestado.status = 'Recusado'
            atestado.observacao_rh = obs
            atestado.save()
            messages.warning(request, f"Atestado de {atestado.funcionario.nome_completo} RECUSADO.")
            
        return redirect('rh_gestao_atestados')

    status_filter = request.GET.get('status', 'Pendente')
    lista = Atestado.objects.all().order_by('-data_envio')
    
    if status_filter != 'Todos':
        lista = lista.filter(status=status_filter)

    return render(request, 'core_rh/rh_gestao_atestados.html', {
        'lista': lista,
        'status_atual': status_filter
    })

@login_required
def admin_atestados_partial_view(request):
    if not (request.user.is_staff or usuario_eh_rh(request.user)):
        return HttpResponse("Acesso Negado", status=403)

    q = request.GET.get('q', '').strip()
    status_filtro = request.GET.get('status', 'Pendente')

    lista = Atestado.objects.all().select_related('funcionario').order_by('-data_envio')

    if q:
        lista = lista.filter(funcionario__nome_completo__icontains=q)

    if status_filtro and status_filtro != 'Todos':
        lista = lista.filter(status=status_filtro)

    context = {
        'lista_atestados': lista,
        'status_atual': status_filtro,
        'q': q
    }
    return render(request, 'core_rh/includes/rh_atestados_moderno.html', context)

@login_required
def rh_acao_atestado(request):
    if not (request.user.is_staff or usuario_eh_rh(request.user)):
        return redirect('home')
        
    if request.method == 'POST':
        atestado_id = request.POST.get('atestado_id')
        acao = request.POST.get('acao')
        obs = request.POST.get('observacao_rh')
        
        atestado = get_object_or_404(Atestado, id=atestado_id)
        
        if acao == 'aprovar':
            atestado.status = 'Aprovado'
            atestado.observacao_rh = obs
            messages.success(request, f"Atestado de {atestado.funcionario.nome_completo} Aprovado!")
        elif acao == 'recusar':
            atestado.status = 'Recusado'
            atestado.observacao_rh = obs
            messages.warning(request, f"Atestado de {atestado.funcionario.nome_completo} Recusado.")
            
        atestado.save()
        
    return redirect(request.META.get('HTTP_REFERER', '/admin/'))

@login_required
def excluir_despesa(request, despesa_id):
    despesa = get_object_or_404(DespesaDiversa, id=despesa_id)
    
    # Segurança opcional: verificar se a despesa pertence ao usuário logado
    # if despesa.funcionario.usuario != request.user:
    #     return HttpResponseForbidden()

    despesa.delete()
    messages.success(request, "Despesa excluída com sucesso.")
    
    # Redireciona de volta para a tela do técnico (ajuste o nome da url se for diferente)
    return redirect('registro_km')

    # --- FUNÇÃO AUXILIAR (Coloque no views.py, fora das views) ---

# --- FUNÇÃO GERADORA DE EXCEL (ATUALIZADA PARA LINK MANUAL) ---
def gerar_workbook_km(funcionario, dt_inicio, dt_fim):
    """Gera o objeto Workbook com os dados de KM/Despesas do funcionário."""
    
    try:
        from pdf2image import convert_from_path
        HAS_PDF_CONVERTER = True
    except ImportError:
        HAS_PDF_CONVERTER = False

    def style_range(ws, cell_range, border=None, fill=None, font=None, alignment=None):
        selection = ws[cell_range]
        if not isinstance(selection, tuple): selection = ((selection,),)
        elif isinstance(selection, tuple) and not isinstance(selection[0], tuple): selection = (selection,)
        for row in selection:
            for cell in row:
                if border: cell.border = border
                if fill: cell.fill = fill
                if font: cell.font = font
                if alignment: cell.alignment = alignment

    lista_itens = []
    val_km = float(funcionario.valor_km) if funcionario.valor_km else 0.0

    # KMs
    kms = ControleKM.objects.filter(funcionario=funcionario, data__range=[dt_inicio, dt_fim]).order_by('data')
    for k in kms:
        trechos = k.trechos.all()
        # Pega a observação salva
        obs_texto = k.observacao if k.observacao else ""
        
        if trechos.exists():
            for t in trechos:
                link = t.origem if t.origem and 'http' in t.origem else None
                origem_final = t.nome_origem if t.nome_origem else "Origem"
                destino_final = t.nome_destino if t.nome_destino else "Destino"
                
                lista_itens.append({
                    'data': k.data, 'chamado': k.numero_chamado, 'tipo': 'DESLOCAMENTO',
                    'origem': origem_final, 'destino': destino_final, 'km': float(t.km),
                    'valor': float(t.km) * val_km, 
                    'obs': obs_texto, # Inclui a observação aqui
                    'link': link,
                    'is_img': False, 'is_pdf': False, 'is_km': True, 'path': None
                })
        else:
            lista_itens.append({
                'data': k.data, 'chamado': k.numero_chamado, 'tipo': 'DESLOCAMENTO',
                'origem': 'Registro Manual', 'destino': '-', 'km': float(k.total_km),
                'valor': float(k.total_km) * val_km, 
                'obs': obs_texto, 
                'link': None,
                'is_img': False, 'is_pdf': False, 'is_km': True, 'path': None
            })

    # Despesas
    despesas = DespesaDiversa.objects.filter(funcionario=funcionario, data__range=[dt_inicio, dt_fim]).order_by('data')
    for d in despesas:
        path_disk = None; url_web = None; is_pdf = False
        if d.comprovante:
            try: url_web = d.comprovante.url 
            except: pass
            try: path_disk = d.comprovante.path
            except:
                try: path_disk = os.path.join(settings.MEDIA_ROOT, d.comprovante.name)
                except: pass
            if path_disk and path_disk.lower().endswith('.pdf'): is_pdf = True
            
        lista_itens.append({
            'data': d.data, 'chamado': d.numero_chamado, 'tipo': d.tipo.upper(),
            'origem': '-', 'destino': '-', 'km': 0.0,
            'valor': float(d.valor or 0), 'obs': d.especificacao or '',
            'link': url_web, 
            'is_img': (path_disk is not None), 'is_pdf': is_pdf, 'is_km': False, 'path': path_disk
        })
    lista_itens.sort(key=lambda x: x['data'])

    # Excel Visual (Layout e Imagens) - MANTIDO IGUAL AO ANTERIOR
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"
    ws.sheet_view.showGridLines = False

    BLUE_DARK = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    BLUE_LIGHT = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    WHITE_FONT = Font(name='Arial', size=14, bold=True, color="FFFFFF")
    BOLD_FONT = Font(name='Arial', size=10, bold=True)
    NORMAL_FONT = Font(name='Arial', size=10)
    BORDER_ALL = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

    ws.column_dimensions['A'].width = 18; ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20; ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 30; ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 15; ws.column_dimensions['H'].width = 25

    # Cabeçalho
    ws.merge_cells('A1:F2'); ws['A1'] = "PLANILHA DE DESPESA"
    style_range(ws, 'A1:F2', border=BORDER_ALL, fill=BLUE_DARK, font=WHITE_FONT, alignment=CENTER)
    ws.merge_cells('G1:G2'); ws['G1'] = "SEMANA"
    style_range(ws, 'G1:G2', border=BORDER_ALL, fill=BLUE_LIGHT, font=BOLD_FONT, alignment=CENTER)
    ws.merge_cells('H1:H2'); ws['H1'] = f"{dt_inicio.isocalendar()[1]}/{dt_inicio.year}"
    style_range(ws, 'H1:H2', border=BORDER_ALL, font=BOLD_FONT, alignment=CENTER)

    labels = [
        (3, 'FUNCIONÁRIO:', funcionario.nome_completo.upper()),
        (4, 'RESIDÊNCIA:', f"{funcionario.endereco} - {funcionario.bairro}"),
        (5, 'TRANSPORTE:', funcionario.tipo_veiculo.upper() if funcionario.tipo_veiculo else "PARTICULAR"),
        (6, 'PERÍODO:', f"DE {dt_inicio.strftime('%d/%m/%Y')} A {dt_fim.strftime('%d/%m/%Y')}")
    ]
    for r, lbl, val in labels:
        ws.cell(row=r, column=1, value=lbl); ws.cell(row=r, column=3, value=val)
        ws.merge_cells(f'A{r}:B{r}'); ws.merge_cells(f'C{r}:F{r}')
        style_range(ws, f'A{r}:B{r}', border=BORDER_ALL, font=BOLD_FONT, alignment=LEFT)
        style_range(ws, f'C{r}:F{r}', border=BORDER_ALL, font=NORMAL_FONT, alignment=LEFT)

    banco_data = [('BANCO', funcionario.banco), ('AGÊNCIA', funcionario.agencia), ('CONTA', funcionario.conta), ('PIX', funcionario.chave_pix)]
    curr_row = 3
    for k, v in banco_data:
        ws.cell(row=curr_row, column=7, value=k); ws.cell(row=curr_row, column=8, value=v or '-')
        style_range(ws, f'G{curr_row}', border=BORDER_ALL, fill=BLUE_LIGHT, font=BOLD_FONT, alignment=CENTER)
        style_range(ws, f'H{curr_row}', border=BORDER_ALL, font=NORMAL_FONT, alignment=CENTER)
        curr_row += 1

    headers_table = ["Data", "Chamado", "Tipo", "Origem / Descrição", "Destino", "KM", "Valor", "Obs / Link"]
    for i, h in enumerate(headers_table, 1):
        ws.cell(row=7, column=i, value=h)
        ws.merge_cells(start_row=7, start_column=i, end_row=8, end_column=i)
        col_letter = get_column_letter(i)
        style_range(ws, f"{col_letter}7:{col_letter}8", border=BORDER_ALL, fill=BLUE_DARK, font=Font(name='Arial', size=11, bold=True, color="FFFFFF"), alignment=CENTER)

    current_row = 9; total_val = 0.0
    wb.create_sheet("Comprovantes"); ws_gal = wb["Comprovantes"]; ws_gal.column_dimensions['A'].width = 80; row_gal = 1

    for item in lista_itens:
        # AQUI: Se for KM, obs já vem com a observação do usuário. Se for Despesa, vem a especificação.
        obs_final = item['obs']
        
        vals = [item['data'].strftime('%d/%m/%Y'), item['chamado'], item['tipo'], item['origem'] if item['is_km'] else obs_final, item['destino'], item['km'] if item['km'] > 0 else "-", item['valor']]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=current_row, column=col, value=val)
            cell.border = BORDER_ALL; cell.font = NORMAL_FONT
            cell.alignment = LEFT if col in [4, 5] else CENTER
            if col == 7: cell.number_format = 'R$ #,##0.00'

        # Coluna OBS/LINK (Coluna 8)
        cell_link = ws.cell(row=current_row, column=8); cell_link.border = BORDER_ALL; cell_link.alignment = CENTER
        
        # Lógica de exibição da coluna 8
        if item['link']:
            txt = "Abrir PDF (Web)" if item['is_pdf'] else ("Abrir Mapa" if item['is_km'] else "Abrir Anexo")
            # Se tiver observação E link, mostra ambos? O Excel só aceita um texto no link.
            # Vamos concatenar se for KM
            if item['is_km'] and obs_final:
                txt = f"{obs_final} (Ver Mapa)"
                
            cell_link.value = txt; cell_link.hyperlink = item['link']; cell_link.font = Font(color="0000FF", underline="single")
        elif item['is_img']:
            cell_link.value = "Ver Aba Comprovantes"; cell_link.font = Font(color="FF0000", italic=True)
        else:
            # Se não tem link nem imagem, mostra apenas a observação (comum em KM manual)
            cell_link.value = obs_final if obs_final else "-"

        if item['path'] and os.path.exists(item['path']):
            ws_gal.cell(row=row_gal, column=1, value=f"REF: {item['data'].strftime('%d/%m')} - R$ {item['valor']} ({item['tipo']})").font = BOLD_FONT
            row_gal += 1
            img_to_insert = None
            try:
                if item['is_pdf'] and HAS_PDF_CONVERTER:
                    images = convert_from_path(item['path'], first_page=1, last_page=1)
                    if images:
                        temp_pdf_img = os.path.join(settings.MEDIA_ROOT, f"temp_pdf_{item['data'].strftime('%d%m%H%M%S')}_{row_gal}.jpg")
                        images[0].save(temp_pdf_img, 'JPEG')
                        img_to_insert = ExcelImage(temp_pdf_img)
                elif not item['is_pdf']:
                    img_to_insert = ExcelImage(item['path'])

                if img_to_insert:
                    base_height = 400; ratio = img_to_insert.width / img_to_insert.height
                    img_to_insert.height = base_height; img_to_insert.width = base_height * ratio
                    ws_gal.add_image(img_to_insert, f'A{row_gal}'); row_gal += 21
                else:
                    ws_gal.cell(row=row_gal, column=1, value="[Conversão indisponível]").font = Font(italic=True); row_gal += 2
            except Exception as e:
                print(f"Erro imagem: {e}"); ws_gal.cell(row=row_gal, column=1, value="[Erro Imagem]").font = Font(color="FF0000"); row_gal += 2

        total_val += item['valor']; current_row += 1

    ws.cell(row=current_row, column=6, value="TOTAL"); ws.cell(row=current_row, column=7, value=total_val)
    style_range(ws, f'F{current_row}:G{current_row}', border=BORDER_ALL, fill=BLUE_LIGHT, font=BOLD_FONT, alignment=CENTER)
    ws.cell(row=current_row, column=7).number_format = 'R$ #,##0.00'
    return wb
# ==============================================================================
# BLOCO KM / DESPESAS / LOTE (Colar no lugar das funções antigas)
# ==============================================================================

# --- DOWNLOAD EM LOTE (NOVO) ---
@login_required
@login_required
def baixar_lote_km(request, equipe_id, ano, mes, semana):
    try:
        # 1. Configuração Básica e Segurança
        equipe = get_object_or_404(Equipe, id=equipe_id)
        
        # 2. Cálculo das Datas da Semana (Robusto)
        try:
            semana_idx = int(semana) - 1
            cal = monthcalendar(int(ano), int(mes))
            semanas_validas = [s for s in cal if any(d != 0 for d in s)]
            
            # Ajuste de limites
            if semana_idx < 0: semana_idx = 0
            if semana_idx >= len(semanas_validas): semana_idx = len(semanas_validas) - 1
            
            semana_lista = semanas_validas[semana_idx]
            dia_referencia = next(d for d in semana_lista if d != 0)
            data_ref = date(int(ano), int(mes), dia_referencia)
            dt_inicio = data_ref - timedelta(days=data_ref.weekday())
            dt_fim = dt_inicio + timedelta(days=6)
        except Exception as e:
            print(f"Erro data: {e}")
            messages.error(request, "Erro ao calcular data da semana.")
            return redirect('area_gestor')

        # 3. Preparação do ZIP na Memória
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            
            # --- A. GERAR PDF DE PAGAMENTO (Na Memória) ---
            pdf_buffer = io.BytesIO()
            doc = SimpleDocTemplate(pdf_buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
            elements = []
            styles = getSampleStyleSheet()

            # Título do PDF
            elements.append(Paragraph(f"Relatório de Pagamento - {equipe.nome}", styles['Title']))
            elements.append(Paragraph(f"Semana {semana}: {dt_inicio.strftime('%d/%m')} a {dt_fim.strftime('%d/%m/%Y')}", styles['Normal']))
            elements.append(Spacer(1, 20))

            # Dados da Tabela
            data_table = [['Técnico', 'Banco', 'Ag/Conta', 'PIX', 'Valor Total (R$)']]
            total_geral = 0.0
            funcionarios = Funcionario.objects.filter(Q(equipe=equipe)|Q(outras_equipes=equipe)).distinct()

            for func in funcionarios:
                kms = ControleKM.objects.filter(funcionario=func, data__range=[dt_inicio, dt_fim]).exclude(status='Rejeitado')
                despesas = DespesaDiversa.objects.filter(funcionario=func, data__range=[dt_inicio, dt_fim]).exclude(status='Rejeitado')
                
                # Adicionar Comprovantes ao ZIP (Se houver)
                for despesa in despesas:
                    if despesa.comprovante:
                        try:
                            ext = despesa.comprovante.name.split('.')[-1]
                            nome_arquivo = f"Comprovantes/{func.nome_completo}_{despesa.tipo_despesa}_{despesa.id}.{ext}"
                            zip_file.writestr(nome_arquivo, despesa.comprovante.read())
                        except:
                            pass # Ignora erro de leitura de arquivo

                # Cálculos Financeiros
                total_km = sum([k.total_km for k in kms])
                total_despesas = sum([d.valor for d in despesas])
                fator = float(func.valor_km) if func.valor_km and func.valor_km > 0 else 1.20
                valor_final = (float(total_km) * fator) + float(total_despesas)

                if valor_final > 0:
                    banco_info = f"{func.banco or '-'} | {func.agencia or '-'} / {func.conta or '-'}"
                    data_table.append([
                        func.nome_completo,
                        func.banco or "-",
                        f"{func.agencia}/{func.conta}",
                        func.chave_pix or "-",
                        f"R$ {valor_final:,.2f}"
                    ])
                    total_geral += valor_final

            data_table.append(['', '', '', 'TOTAL:', f"R$ {total_geral:,.2f}"])

            # Estilo Tabela PDF
            t = Table(data_table, colWidths=[200, 100, 150, 150, 100])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, -1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(t)
            
            # Finaliza PDF
            doc.build(elements)
            
            # Escreve o PDF no ZIP
            nome_filial = equipe.nome.replace('Campo ', '').strip()
            zip_file.writestr(f"Relatorio_{nome_filial}_Semana_{semana}.pdf", pdf_buffer.getvalue())

        # 4. Retorna o ZIP
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="Lote_{equipe.nome}_S{semana}.zip"'
        return response

    except Exception as e:
        # Log do erro no console para debug
        print(f"ERRO 500 no ZIP: {str(e)}")
        # Se falhar tudo, retorna erro simples para não quebrar a página
        return HttpResponse(f"Erro ao gerar lote: {str(e)}", status=500)
@login_required
def aprovar_semana_lote(request, equipe_id, ano, mes, semana):
    from datetime import date, timedelta
    from calendar import monthcalendar
    
    # 1. Cálculo de Datas
    try:
        semana_idx = int(semana) - 1
        cal = monthcalendar(int(ano), int(mes))
        semanas_validas = [s for s in cal if any(d != 0 for d in s)]
        if semana_idx < 0: semana_idx = 0
        if semana_idx >= len(semanas_validas): semana_idx = len(semanas_validas) - 1
        
        semana_lista = semanas_validas[semana_idx]
        dia_referencia = next(d for d in semana_lista if d != 0)
        data_ref = date(int(ano), int(mes), dia_referencia)
        dt_inicio = data_ref - timedelta(days=data_ref.weekday())
        dt_fim = dt_inicio + timedelta(days=6)
    except:
        messages.error(request, "Data inválida.")
        return redirect('area_gestor')

    equipe = get_object_or_404(Equipe, id=equipe_id)
    funcionarios = Funcionario.objects.filter(Q(equipe=equipe)|Q(outras_equipes=equipe)).distinct()

    kms_qs = ControleKM.objects.filter(funcionario__in=funcionarios, data__range=[dt_inicio, dt_fim])
    desp_qs = DespesaDiversa.objects.filter(funcionario__in=funcionarios, data__range=[dt_inicio, dt_fim])

    count = 0
    etapa = ""

    # LÓGICA ESTRITA DE HIERARQUIA
    # O lote só move o que está na "caixa de entrada" daquela função específica

    if usuario_eh_financeiro(request.user):
        # Financeiro tem 2 passos: Aprovar o que veio da Matriz E Pagar o que ele já aprovou
        # 1. Aprovado_Matriz -> Aprovado_Financeiro
        k1 = kms_qs.filter(status='Aprovado_Matriz').update(status='Aprovado_Financeiro')
        d1 = desp_qs.filter(status='Aprovado_Matriz').update(status='Aprovado_Financeiro')
        
        # 2. Aprovado_Financeiro -> Pago
        k2 = kms_qs.filter(status='Aprovado_Financeiro').update(status='Pago')
        d2 = desp_qs.filter(status='Aprovado_Financeiro').update(status='Pago')
        
        count = k1 + d1 + k2 + d2
        etapa = "Financeiro"

    elif usuario_eh_gestao(request.user):
        # Matriz SÓ aprova o que já é 'Aprovado_Regional'
        # Não toca em 'Pendente'
        k = kms_qs.filter(status='Aprovado_Regional').update(status='Aprovado_Matriz')
        d = desp_qs.filter(status='Aprovado_Regional').update(status='Aprovado_Matriz')
        count = k + d
        etapa = "Matriz"

    else:
        # Regional (Gestor Comum) SÓ aprova o que é 'Pendente'
        k = kms_qs.filter(status='Pendente').update(status='Aprovado_Regional')
        d = desp_qs.filter(status='Pendente').update(status='Aprovado_Regional')
        count = k + d
        etapa = "Regional"
    
    if count > 0:
        messages.success(request, f"Lote ({etapa}) processado: {count} itens avançaram de etapa.")
    else:
        messages.info(request, f"Nenhum item aguardando a etapa ({etapa}) nesta semana.")
        
    return redirect('area_gestor')

# --- REGISTRO KM E DOWNLOAD INDIVIDUAL (SEM SELENIUM - 100% MANUAL) ---
@login_required
def registro_km_view(request):
    if not usuario_eh_campo(request.user): return redirect('home')
    funcionario = request.user.funcionario

    if request.method == 'POST':
        # Pega as listas de dados
        datas = request.POST.getlist('data_viagem[]')
        chamados = request.POST.getlist('numero_chamado[]')
        origens = request.POST.getlist('nome_origem[]')
        destinos = request.POST.getlist('nome_destino[]')
        kms_lista = request.POST.getlist('km_manual[]')
        urls = request.POST.getlist('google_url[]')
        observacoes = request.POST.getlist('observacao[]') # NOVO CAMPO
        
        saved_count = 0
        
        for i in range(len(datas)):
            try:
                if not datas[i] or not kms_lista[i]:
                    continue

                data_final = datetime.strptime(datas[i], '%Y-%m-%d').date()
                
                if is_periodo_travado(funcionario, data_final):
                    continue

                km_final = float(kms_lista[i].replace(',', '.'))
                obs_final = observacoes[i] if i < len(observacoes) else "" # Pega obs correspondente
                
                if km_final > 0:
                    with transaction.atomic():
                        controle = ControleKM.objects.create(
                            funcionario=funcionario, 
                            data=data_final, 
                            total_km=km_final, 
                            numero_chamado=chamados[i], 
                            observacao=obs_final, # SALVA A OBS
                            status='Pendente'
                        )
                        TrechoKM.objects.create(
                            controle=controle, 
                            origem=urls[i] if urls[i] else "-", 
                            destino='-', 
                            km=km_final,
                            nome_origem=origens[i], 
                            nome_destino=destinos[i]
                        )
                    saved_count += 1
            except Exception as e:
                print(f"Erro ao salvar item {i}: {e}")
                continue

        if saved_count > 0:
            messages.success(request, f"{saved_count} rotas registradas com sucesso!")
        
        return redirect('registro_km')
    
    # GET: Carrega histórico
    kms = ControleKM.objects.filter(funcionario=funcionario).order_by('-data')[:50]
    despesas = DespesaDiversa.objects.filter(funcionario=funcionario).order_by('-data')[:50]
    
    historico = []
    for k in kms:
        historico.append({
            'id': k.id, 'data': k.data, 'numero_chamado': k.numero_chamado, 
            'is_km': True, 'trechos': k.trechos.all(), 'total_km': k.total_km, 
            'valor': None, 'status': k.status, 'nota_recusa': k.nota_recusa,
            'observacao': k.observacao # Passa para o template
        })
    for d in despesas:
        historico.append({
            'id': d.id, 'data': d.data, 'numero_chamado': d.numero_chamado, 
            'is_km': False, 'tipo_despesa': d.tipo, 'valor': d.valor, 
            'status': d.status, 'comprovante': d.comprovante, 
            'especificacao': d.especificacao, 'nota_recusa': d.nota_recusa
        })
    
    historico.sort(key=lambda x: x['data'], reverse=True)
    return render(request, 'core_rh/registro_km.html', {'historico': historico, 'funcionario': funcionario})
@login_required
# --- DOWNLOAD INDIVIDUAL EXCEL (ATUALIZADO) ---
@login_required
def baixar_relatorio_excel(request, func_id=None):
    if func_id and (request.user.is_staff or usuario_eh_rh(request.user)):
        funcionario = get_object_or_404(Funcionario, id=func_id)
    else:
        try: funcionario = request.user.funcionario
        except: return HttpResponse("Erro.", status=400)

    hoje = timezone.now().date()
    try:
        ano = int(request.GET.get('ano', hoje.year))
        mes = int(request.GET.get('mes', hoje.month))
        semana_param = request.GET.get('semana')
    except: ano, mes, semana_param = hoje.year, hoje.month, None

    # Calcula datas (Lógica mantida)
    if semana_param:
        try:
            semana_idx = int(semana_param) - 1
            cal = monthcalendar(ano, mes)
            semanas_validas = [s for s in cal if any(d != 0 for d in s)]
            if semana_idx < 0: semana_idx = 0
            if semana_idx >= len(semanas_validas): semana_idx = len(semanas_validas) - 1
            
            semana_lista = semanas_validas[semana_idx]
            dia_referencia = next(d for d in semana_lista if d != 0)
            data_ref = date(ano, mes, dia_referencia)
            
            dt_inicio = data_ref - timedelta(days=data_ref.weekday())
            dt_fim = dt_inicio + timedelta(days=6)
        except: dt_inicio = date(ano, mes, 1); dt_fim = date(ano, mes, monthrange(ano, mes)[1])
    else:
        dt_inicio = date(ano, mes, 1); dt_fim = date(ano, mes, monthrange(ano, mes)[1])

    wb = gerar_workbook_km(funcionario, dt_inicio, dt_fim)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    # --- NOVA LÓGICA DE NOME DA FILIAL ---
    # Prioridade: Equipe Secundária que tenha "Campo" no nome
    equipe_filial = funcionario.outras_equipes.filter(nome__icontains="Campo").first()
    
    if equipe_filial:
        raw_name = equipe_filial.nome
    elif funcionario.equipe:
        raw_name = funcionario.equipe.nome
    else:
        raw_name = "Geral"
        
    # Limpa o nome: Remove "Campo ", "campo " e retira espaços
    nome_filial = raw_name.replace('Campo ', '').replace('campo ', '').strip().replace(' ', '')
    
    # --- FORMATAÇÃO DO NOME DO ARQUIVO ---
    # 1. Nome e Sobrenome
    partes = funcionario.nome_completo.split()
    nome_func = f"{partes[0]}{partes[1]}" if len(partes) > 1 else partes[0]
    
    # 2. Semana do Ano
    semana_anual = dt_inicio.isocalendar()[1]
    
    # Ex: LucasAntonio_KM_SaoPaulo_S2.xlsx
    nome_arq = f"{nome_func}_KM_{nome_filial}_S{semana_anual}.xlsx"
    
    response['Content-Disposition'] = f'attachment; filename="{nome_arq}"'
    wb.save(response)
    return response
@login_required
def avancar_status_km(request, controle_id):
    km = get_object_or_404(ControleKM, id=controle_id)
    user = request.user
    status_atual = km.status
    novo_status = None
    msg_sucesso = ""

    # REINICIAR (Se estiver rejeitado ou vazio/bugado)
    if status_atual == 'Rejeitado' or status_atual == '':
        novo_status = 'Pendente'
        msg_sucesso = "Registro reiniciado para Pendente."

    # 1. Pendente -> Aprovado Regional
    elif status_atual == 'Pendente':
        novo_status = 'Aprovado_Regional'
        msg_sucesso = "Aprovação Regional realizada."

    # 2. Regional -> Matriz
    elif status_atual == 'Aprovado_Regional': # Verifica se o texto está exato
        if usuario_eh_gestao(user):
            novo_status = 'Aprovado_Matriz'
            msg_sucesso = "Aprovação da Matriz realizada."
        else:
            messages.error(request, "Permissão negada: Necessário equipe Gestão/Matriz.")
            return redirect('area_gestor')

    # 3. Matriz -> Financeiro
    elif status_atual == 'Aprovado_Matriz':
        if usuario_eh_financeiro(user):
            novo_status = 'Aprovado_Financeiro'
            msg_sucesso = "Aprovação Financeira realizada."
        else:
            messages.error(request, "Permissão negada: Necessário equipe Financeiro.")
            return redirect('area_gestor')

    # 4. Financeiro -> Pago
    elif status_atual == 'Aprovado_Financeiro':
        if usuario_eh_financeiro(user):
            novo_status = 'Pago'
            msg_sucesso = "Pagamento confirmado."
        else:
            messages.error(request, "Permissão negada.")
            return redirect('area_gestor')

    if novo_status:
        # Atualiza
        km.status = novo_status
        km.save()
        
        # Sincroniza semana
        dt = km.data
        ini = dt - timedelta(days=dt.weekday())
        fim = ini + timedelta(days=6)
        
        DespesaDiversa.objects.filter(funcionario=km.funcionario, data__range=[ini, fim]).update(status=novo_status)
        ControleKM.objects.filter(funcionario=km.funcionario, data__range=[ini, fim]).update(status=novo_status)
        
        messages.success(request, msg_sucesso)
    else:
        # Se caiu aqui, é porque o status no banco não bate com a lógica
        messages.warning(request, f"Status desconhecido ou permissão insuficiente. Status atual: {status_atual}")

    return redirect('area_gestor')
@login_required
def rejeitar_km_gestor(request, controle_id):
    if request.method == 'POST':
        km = get_object_or_404(ControleKM, id=controle_id)
        
        motivo = request.POST.get('motivo_recusa')
        
        # Define status REJEITADO (para aparecer vermelho pro técnico)
        novo_status = 'Rejeitado'
        
        dt = km.data
        ini = dt - timedelta(days=dt.weekday())
        fim = ini + timedelta(days=6)

        # Atualiza Despesas
        DespesaDiversa.objects.filter(
            funcionario=km.funcionario, 
            data__range=[ini, fim]
        ).update(status=novo_status, nota_recusa=motivo)
        
        # Atualiza KMs
        ControleKM.objects.filter(
            funcionario=km.funcionario, 
            data__range=[ini, fim]
        ).update(status=novo_status, nota_recusa=motivo)
        
        messages.warning(request, "Registro rejeitado. O técnico foi notificado.")
        
    return redirect('area_gestor')
@login_required
def excluir_km(request, km_id):
    c = get_object_or_404(ControleKM, id=km_id)
    if c.funcionario.usuario == request.user or request.user.is_superuser:
        # --- TRAVA DE EXCLUSÃO ---
        if c.status in ['Aprovado', 'Pago']:
            messages.error(request, "Não é possível excluir um registro já Aprovado ou Pago.")
        else:
            c.delete()
            messages.success(request, "Removido")
    return redirect('registro_km')

@login_required
def excluir_despesa(request, despesa_id):
    d = get_object_or_404(DespesaDiversa, id=despesa_id)
    if d.funcionario.usuario == request.user or request.user.is_superuser:
        # --- TRAVA DE EXCLUSÃO ---
        if d.status in ['Aprovado', 'Pago']:
            messages.error(request, "Não é possível excluir uma despesa já Aprovada ou Paga.")
        else:
            d.delete()
            messages.success(request, "Despesa Removida")
    return redirect('registro_km')

@login_required
def salvar_despesa_diversa_view(request):
    if request.method != 'POST': return redirect('registro_km')
    try: funcionario = request.user.funcionario
    except: return redirect('registro_km')

    datas = request.POST.getlist('data_despesa[]')
    chamados = request.POST.getlist('numero_chamado[]')
    tipos = request.POST.getlist('tipo_despesa[]')
    especs = request.POST.getlist('especificacao[]')
    valores = request.POST.getlist('valor[]')
    arquivos = request.FILES.getlist('comprovante[]') # Atenção: Ordem dos arquivos pode ser tricky no HTML

    saved_count = 0
    
    # OBS: O input file multiple ou lista de inputs file requer cuidado no frontend.
    # Assumindo que o HTML enviará 'comprovante[]' corretamente indexado.
    
    for i in range(len(datas)):
        try:
            if not datas[i] or not valores[i]: continue
            
            data_final = datetime.strptime(datas[i], '%Y-%m-%d').date()
            
            # --- TRAVA DE SEGURANÇA ---
            if is_periodo_travado(funcionario, data_final):
                continue
            
            valor_str = valores[i]
            valor_final = 0.00
            if valor_str:
                limpo = valor_str.replace('R$', '').replace(' ', '').replace('.', '').replace(',', '.')
                try: valor_final = float(limpo)
                except: valor_final = 0.00

            # Pega arquivo correspondente (se existir)
            arquivo = arquivos[i] if i < len(arquivos) else None
            
            if not arquivo: continue # Obrigatório

            DespesaDiversa.objects.create(
                funcionario=funcionario, data=data_final, numero_chamado=chamados[i],
                tipo=tipos[i], especificacao=especs[i], valor=valor_final,
                comprovante=arquivo, status='Pendente'
            )
            saved_count += 1
        except Exception as e:
            print(f"Erro despesa {i}: {e}")
            continue

    if saved_count > 0:
        messages.success(request, f"{saved_count} despesas lançadas!")
    else:
        messages.error(request, "Erro ao salvar despesas. Verifique os dados.")

    return redirect('registro_km')
@login_required
def atualizar_dados_tecnico(request):
    if request.method == 'POST':
        try:
            try: func = request.user.funcionario
            except: return redirect('registro_km')
            
            func.cep = request.POST.get('cep')
            func.endereco = request.POST.get('endereco')
            func.bairro = request.POST.get('bairro')
            func.cidade = request.POST.get('cidade')
            func.estado = request.POST.get('estado')
            func.base = request.POST.get('base')
            func.tipo_veiculo = request.POST.get('tipo_veiculo')
            
            val_km = request.POST.get('valor_km', '0')
            if val_km:
                val_limpo = val_km.replace('R$', '').replace(' ', '').replace('.', '').replace(',', '.')
                try: func.valor_km = float(val_limpo)
                except: pass
            
            func.banco = request.POST.get('banco')
            func.agencia = request.POST.get('agencia')
            func.operacao = request.POST.get('operacao')
            func.conta = request.POST.get('conta')
            
            # --- CORREÇÃO: SALVANDO O PIX ---
            func.chave_pix = request.POST.get('chave_pix')
            
            func.save()
            messages.success(request, "Dados atualizados!")
        except Exception as e:
            messages.error(request, f"Erro: {e}")
            
    return redirect('registro_km')
@login_required
def repetir_rota_view(request):
    if request.method == 'POST':
        try:
            km_id = request.POST.get('km_id')
            nova_data = request.POST.get('nova_data')
            novo_chamado = request.POST.get('novo_chamado')
            
            # --- TRAVA DE SEGURANÇA ---
            if is_periodo_travado(request.user.funcionario, nova_data):
                messages.error(request, "ERRO: A data escolhida pertence a uma semana já fechada/aprovada.")
                return redirect('registro_km')
            # --------------------------

            original_controle = get_object_or_404(ControleKM, id=km_id)
            if original_controle.funcionario.usuario != request.user:
                return redirect('registro_km')

            novo_controle = ControleKM.objects.create(
                funcionario=request.user.funcionario,
                data=nova_data,
                numero_chamado=novo_chamado,
                total_km=original_controle.total_km,
                status='Pendente'
            )

            trecho_original = original_controle.trechos.first()
            if trecho_original:
                TrechoKM.objects.create(
                    controle=novo_controle,
                    origem=trecho_original.origem,
                    destino=trecho_original.destino,
                    km=trecho_original.km,
                    nome_origem=trecho_original.nome_origem,
                    nome_destino=trecho_original.nome_destino
                )
            messages.success(request, "Rota repetida com sucesso!")
        except Exception as e:
            messages.error(request, f"Erro: {e}")

    return redirect('registro_km')
def is_periodo_travado(funcionario, data_ref):
    """
    Retorna True se a semana da data_ref já tiver itens em processo de aprovação ou pagos.
    Bloqueia edição para manter a integridade do relatório do gestor.
    """
    if isinstance(data_ref, str):
        try: data_ref = datetime.strptime(data_ref, '%Y-%m-%d').date()
        except: return False 
        
    start = data_ref - timedelta(days=data_ref.weekday()) # Segunda
    end = start + timedelta(days=6) # Domingo
    
    # Lista de status que bloqueiam a edição (qualquer coisa que não seja Pendente ou Rejeitado)
    status_bloqueio = [
        'Aprovado_Regional', 
        'Aprovado_Matriz', 
        'Aprovado_Financeiro', 
        'Pago', 
        'Aprovado' # Mantido para compatibilidade com dados antigos
    ]
    
    bloq_km = ControleKM.objects.filter(
        funcionario=funcionario, 
        data__range=[start, end], 
        status__in=status_bloqueio
    ).exists()
    
    bloq_desp = DespesaDiversa.objects.filter(
        funcionario=funcionario, 
        data__range=[start, end], 
        status__in=status_bloqueio
    ).exists()
    
    return bloq_km or bloq_desp
def usuario_eh_gestao(user):
    """Verifica se o usuário está na equipe 'Gestão' ou 'Matriz' (Ignora acentos/case)."""
    if not user.is_authenticated: return False
    if user.is_superuser: return True
    
    try:
        func = user.funcionario
        # Lista de nomes aceitos (Adicione variações aqui)
        nomes_permitidos = ['Gestão', 'Gestao', 'Matriz', 'Diretoria', 'Administrativo']
        
        # Verifica Equipe Principal
        if func.equipe and func.equipe.nome in nomes_permitidos:
            return True
            
        # Verifica Equipes Secundárias
        if func.outras_equipes.filter(nome__in=nomes_permitidos).exists():
            return True
    except:
        pass
    return False

def usuario_eh_financeiro(user):
    """Verifica se o usuário está na equipe 'Financeiro'."""
    if not user.is_authenticated: return False
    if user.is_superuser: return True
    
    try:
        func = user.funcionario
        nomes_permitidos = ['Financeiro', 'Financeira', 'Finanças']
        
        if func.equipe and func.equipe.nome in nomes_permitidos:
            return True
        if func.outras_equipes.filter(nome__in=nomes_permitidos).exists():
            return True
    except:
        pass
    return False

@login_required
def resetar_status_bugados(request):
    if not request.user.is_superuser: return redirect('home')
    
    # Reseta tudo que não for um status válido padrão
    status_validos = ['Pendente', 'Aprovado_Regional', 'Aprovado_Matriz', 'Aprovado_Financeiro', 'Pago', 'Rejeitado']
    
    qtd = ControleKM.objects.exclude(status__in=status_validos).update(status='Pendente')
    DespesaDiversa.objects.exclude(status__in=status_validos).update(status='Pendente')
    
    messages.success(request, f"{qtd} registros corrompidos foram resetados para Pendente.")
    return redirect('area_gestor')

# Certifique-se de ter estes imports no topo



from collections import defaultdict
from datetime import datetime

@login_required
def gerar_relatorio_customizado(request):
    if request.method != 'POST':
        return redirect('area_gestor')

    # 1. VALIDAÇÃO E FILTROS
    try:
        data_inicio = datetime.strptime(request.POST.get('data_inicio'), '%Y-%m-%d').date()
        data_fim = datetime.strptime(request.POST.get('data_fim'), '%Y-%m-%d').date()
        equipes_ids = request.POST.getlist('equipes')
    except (ValueError, TypeError):
        messages.error(request, "Dados inválidos.")
        return redirect('area_gestor')

    if not equipes_ids:
        messages.error(request, "Selecione ao menos uma equipe.")
        return redirect('area_gestor')

    # 2. QUERIES
    equipes = Equipe.objects.filter(id__in=equipes_ids)
    funcionarios = Funcionario.objects.filter(Q(equipe__in=equipes) | Q(outras_equipes__in=equipes)).distinct()

    kms = ControleKM.objects.filter(funcionario__in=funcionarios, data__range=[data_inicio, data_fim]).exclude(status='Rejeitado').select_related('funcionario', 'funcionario__equipe')
    despesas = DespesaDiversa.objects.filter(funcionario__in=funcionarios, data__range=[data_inicio, data_fim]).exclude(status='Rejeitado').select_related('funcionario', 'funcionario__equipe')

    # 3. ESTRUTURAS DE DADOS
    
    # Estrutura GLOBAL
    global_data = defaultdict(lambda: {
        'total_periodo': 0.0,
        'mensal': defaultdict(float),
        'semanal': defaultdict(float),
        'diario': defaultdict(float)
    })

    # Estrutura DETALHADA POR FILIAL (Antiga Equipe)
    team_data = defaultdict(lambda: defaultdict(lambda: {
        'total_periodo': 0.0,
        'tipos': defaultdict(float),
        'mensal': defaultdict(lambda: defaultdict(float)),
        'semanal': defaultdict(lambda: defaultdict(float)),
        'diario': defaultdict(lambda: defaultdict(float))
    }))

    all_months = set()
    all_weeks = set()
    all_days = set()
    all_types = set()

    # --- FUNÇÃO DE LIMPEZA DO NOME (FILIAL) ---
    def get_filial_name(func):
        raw_name = "Indefinida"
        if func.equipe and str(func.equipe.id) in equipes_ids: 
            raw_name = func.equipe.nome
        else:
            for eq in func.outras_equipes.all():
                if str(eq.id) in equipes_ids: 
                    raw_name = eq.nome
                    break
        
        # Remove "Campo " (com o espaço) e retorna
        # Ex: "Campo CIAUSRE" -> "CIAUSRE"
        return raw_name.replace("Campo ", "").replace("campo ", "")

    # --- PROCESSA KMs ---
    for k in kms:
        func = k.funcionario
        filial = get_filial_name(func) # Nome já limpo
        nome = func.nome_completo
        val = float(k.total_km) * (float(func.valor_km) if func.valor_km else 0.0)
        
        mes = k.data.strftime('%m/%Y')
        sem = f"{k.data.isocalendar()[1]}/{k.data.year}"
        dia = k.data.strftime('%d/%m/%Y')

        # Popula GLOBAL
        global_data[filial]['total_periodo'] += val
        global_data[filial]['mensal'][mes] += val
        global_data[filial]['semanal'][sem] += val
        global_data[filial]['diario'][dia] += val

        # Popula DETALHADO
        ref = team_data[filial][nome]
        ref['total_periodo'] += val
        ref['tipos']['KM'] += val
        ref['mensal'][mes]['KM'] += val
        ref['semanal'][sem]['KM'] += val
        ref['diario'][dia]['KM'] += val

        all_months.add(mes)
        all_weeks.add(sem)
        all_days.add(dia)
        all_types.add('KM')

    # --- PROCESSA DESPESAS ---
    for d in despesas:
        func = d.funcionario
        filial = get_filial_name(func) # Nome já limpo
        nome = func.nome_completo
        val = float(d.valor)
        tipo = d.tipo 
        
        mes = d.data.strftime('%m/%Y')
        sem = f"{d.data.isocalendar()[1]}/{d.data.year}"
        dia = d.data.strftime('%d/%m/%Y')

        # Popula GLOBAL
        global_data[filial]['total_periodo'] += val
        global_data[filial]['mensal'][mes] += val
        global_data[filial]['semanal'][sem] += val
        global_data[filial]['diario'][dia] += val

        # Popula DETALHADO
        ref = team_data[filial][nome]
        ref['total_periodo'] += val
        ref['tipos'][tipo] += val
        ref['mensal'][mes][tipo] += val
        ref['semanal'][sem][tipo] += val
        ref['diario'][dia][tipo] += val

        all_months.add(mes)
        all_weeks.add(sem)
        all_days.add(dia)
        all_types.add(tipo)

    # Ordenações
    sorted_months = sorted(list(all_months), key=lambda x: datetime.strptime(x, '%m/%Y'))
    sorted_days = sorted(list(all_days), key=lambda x: datetime.strptime(x, '%d/%m/%Y'))
    sorted_weeks = sorted(list(all_weeks), key=lambda x: (int(x.split('/')[1]), int(x.split('/')[0])))
    
    # Ordena tipos garantindo KM primeiro
    temp_types = list(all_types)
    if 'KM' in temp_types:
        temp_types.remove('KM')
        sorted_types = ['KM'] + sorted(temp_types)
    else:
        sorted_types = sorted(temp_types)

    # 4. EXCEL GENERATION
    wb = Workbook()
    
    # Estilos
    header_style = NamedStyle(name="header")
    header_style.font = Font(bold=True, color="FFFFFF")
    header_style.fill = PatternFill("solid", fgColor="4F81BD")
    header_style.alignment = Alignment(horizontal="center", vertical="center")
    
    period_header_style = NamedStyle(name="period_header")
    period_header_style.font = Font(bold=True)
    period_header_style.fill = PatternFill("solid", fgColor="DCE6F1")
    period_header_style.alignment = Alignment(horizontal="center")
    
    money_fmt = 'R$ #,##0.00'
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Função de ajuste de largura
    def auto_adjust_width(ws):
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        cell_len = len(str(cell.value))
                        if cell.font and cell.font.bold: cell_len *= 1.2
                        if cell_len > max_length: max_length = cell_len
                except: pass
            adjusted_width = (max_length + 2) * 1.1
            if adjusted_width < 12: adjusted_width = 12
            if adjusted_width > 60: adjusted_width = 60
            ws.column_dimensions[column_letter].width = adjusted_width

    def write_header(ws, row, cols, col_offset=1):
        for i, c in enumerate(cols):
            cell = ws.cell(row=row, column=i+col_offset, value=c)
            cell.style = header_style
            cell.border = border

    def write_row(ws, row, cols, bold_last=False):
        for i, val in enumerate(cols, 1):
            cell = ws.cell(row=row, column=i, value=val)
            cell.border = border
            if isinstance(val, (int, float)):
                cell.number_format = money_fmt
            if bold_last and i == len(cols):
                cell.font = Font(bold=True)

    # === 1. ABAS GLOBAIS (AGRUPADO POR FILIAL) ===

    def create_global_sheet(title, time_cols, data_key):
        if title == "Resumo Geral": ws = wb.active; ws.title = title
        else: ws = wb.create_sheet(title)
        
        # MUDANÇA AQUI: "Equipe" vira "Filial"
        headers = ["Filial"] + (time_cols if time_cols else ["TOTAL (R$)"]) 
        if time_cols: headers.append("TOTAL")
        
        write_header(ws, 1, headers)
        
        row = 2
        for filial in sorted(global_data.keys()):
            vals = [filial]
            if not time_cols:
                vals.append(global_data[filial]['total_periodo'])
            else:
                row_total = 0
                for t in time_cols:
                    v = global_data[filial][data_key].get(t, 0.0)
                    vals.append(v)
                    row_total += v
                vals.append(row_total)
            
            write_row(ws, row, vals, bold_last=True)
            row += 1
        
        auto_adjust_width(ws)

    # Gera as 4 abas globais
    create_global_sheet("Resumo Geral", [], None)
    create_global_sheet("Visão Mensal (Global)", sorted_months, 'mensal')
    create_global_sheet("Visão Semanal (Global)", sorted_weeks, 'semanal')
    create_global_sheet("Visão Diária (Global)", sorted_days, 'diario')

    # === 2. ABAS POR FILIAL (DETALHADO POR FUNCIONÁRIO E TIPO) ===
    
    for filial in sorted(team_data.keys()):
        # Nome da aba seguro
        safe_name = filial[:30].replace('/', '-')
        ws = wb.create_sheet(safe_name)
        
        current_row = 1
        # Título interno
        ws.cell(row=current_row, column=1, value=f"RELATÓRIO: {filial.upper()}").font = Font(bold=True, size=14)
        current_row += 2

        def add_complex_table(title, time_periods, data_key):
            nonlocal current_row
            ws.cell(row=current_row, column=1, value=title.upper()).font = Font(bold=True, size=12, color="4F81BD")
            current_row += 1
            col_idx = 2
            ws.cell(row=current_row, column=1, value="Colaborador").style = header_style
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row+1, end_column=1)
            
            # Linha 1: Períodos
            for period in time_periods:
                start_col = col_idx
                end_col = col_idx + len(sorted_types) - 1 
                if start_col < end_col:
                    ws.merge_cells(start_row=current_row, start_column=start_col, end_row=current_row, end_column=end_col)
                cell = ws.cell(row=current_row, column=start_col, value=period)
                cell.style = period_header_style
                cell.border = border
                col_idx = end_col + 1
            
            # Total Geral
            ws.merge_cells(start_row=current_row, start_column=col_idx, end_row=current_row+1, end_column=col_idx)
            cell = ws.cell(row=current_row, column=col_idx, value="TOTAL GERAL")
            cell.style = header_style
            total_geral_col_idx = col_idx
            current_row += 1
            
            # Linha 2: Tipos
            col_idx = 2
            for period in time_periods:
                for tipo in sorted_types:
                    cell = ws.cell(row=current_row, column=col_idx, value=tipo)
                    cell.style = header_style
                    cell.font = Font(bold=True, size=9, color="FFFFFF")
                    col_idx += 1
            current_row += 1
            
            # Dados
            for func in sorted(team_data[filial].keys()):
                dados_func = team_data[filial][func]
                ws.cell(row=current_row, column=1, value=func).border = border
                col_idx = 2
                func_grand_total = 0
                for period in time_periods:
                    period_values = dados_func[data_key].get(period, {})
                    for tipo in sorted_types:
                        val = period_values.get(tipo, 0.0)
                        cell = ws.cell(row=current_row, column=col_idx, value=val)
                        cell.number_format = money_fmt
                        cell.border = border
                        func_grand_total += val
                        col_idx += 1
                cell = ws.cell(row=current_row, column=total_geral_col_idx, value=func_grand_total)
                cell.number_format = money_fmt
                cell.font = Font(bold=True)
                cell.border = border
                current_row += 1
            current_row += 2

        # 1. Resumo Simples por Tipo
        ws.cell(row=current_row, column=1, value="RESUMO POR TIPO").font = Font(bold=True, color="4F81BD")
        current_row += 1
        headers_tipo = ["Colaborador"] + sorted_types + ["TOTAL"]
        write_header(ws, current_row, headers_tipo, col_offset=1)
        current_row += 1
        for func in sorted(team_data[filial].keys()):
            vals = [func]
            tot = 0
            for t in sorted_types:
                v = team_data[filial][func]['tipos'].get(t, 0.0)
                vals.append(v); tot += v
            vals.append(tot)
            write_row(ws, current_row, vals, bold_last=True)
            current_row += 1
        current_row += 2

        # 2. Tabelas Complexas
        add_complex_table("Detalhamento Mensal", sorted_months, 'mensal')
        add_complex_table("Detalhamento Semanal", sorted_weeks, 'semanal')
        add_complex_table("Detalhamento Diário", sorted_days, 'diario')

        auto_adjust_width(ws)

    # Finaliza
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    nome_arq = f"Relatorio_Filiais_{data_inicio.strftime('%d%m')}_{data_fim.strftime('%d%m')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{nome_arq}"'
    wb.save(response)
    
    return response


class CustomPasswordResetView(PasswordResetView):
    form_class = CpfPasswordResetForm
    template_name = 'registration/password_reset_form.html'
    
    # Adicione estas duas linhas:
    email_template_name = 'registration/password_reset_email.html' # Versão Texto
    html_email_template_name = 'registration/password_reset_email_html.html' # Versão HTML (O Bonito)
    
    success_url = reverse_lazy('password_reset_done')

    def form_valid(self, form):
        email = form.cleaned_data.get('email')
        if email:
            self.request.session['reset_email'] = email
        return super().form_valid(form)
class CustomPasswordResetDoneView(PasswordResetDoneView):
    template_name = 'registration/password_reset_done.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # RECUPERA O E-MAIL DA SESSÃO E MASCARA
        email = self.request.session.get('reset_email')
        if email:
            try:
                user_part, domain_part = email.split('@')
                visible = user_part[:3]
                masked_email = f"{visible}***@{domain_part}"
                context['masked_email'] = masked_email
            except ValueError:
                context['masked_email'] = email
        return context
@login_required
def editar_km_view(request):
    if request.method == 'POST':
        km_id = request.POST.get('km_id')
        try:
            km = ControleKM.objects.get(id=km_id, funcionario__usuario=request.user)
            
            if km.status not in ['Pendente', 'Rejeitado']:
                messages.error(request, "Não é possível editar este registro (Status bloqueado).")
                return redirect('registro_km')
            
            km.data = request.POST.get('data_viagem')
            km.numero_chamado = request.POST.get('numero_chamado')
            km.total_km = float(request.POST.get('km_manual').replace(',', '.'))
            km.observacao = request.POST.get('observacao') # EDITA OBS
            
            km.status = 'Pendente'
            km.nota_recusa = None 
            km.save()
            
            trecho = km.trechos.first()
            if trecho:
                trecho.nome_origem = request.POST.get('nome_origem')
                trecho.nome_destino = request.POST.get('nome_destino')
                trecho.origem = request.POST.get('google_url')
                trecho.km = km.total_km
                trecho.save()
                
            messages.success(request, "Registro corrigido e reenviado!")
            
        except Exception as e:
            messages.error(request, f"Erro ao editar: {e}")
            
    return redirect('registro_km')

@login_required
def gerar_pdf_pagamento_equipe(request, equipe_id, ano, mes, semana):
    # 1. Calcular datas da semana
    try:
        semana_idx = int(semana) - 1
        cal = monthcalendar(int(ano), int(mes))
        semanas_validas = [s for s in cal if any(d != 0 for d in s)]
        if semana_idx < 0: semana_idx = 0
        if semana_idx >= len(semanas_validas): semana_idx = len(semanas_validas) - 1
        
        semana_lista = semanas_validas[semana_idx]
        dia_referencia = next(d for d in semana_lista if d != 0)
        data_ref = date(int(ano), int(mes), dia_referencia)
        dt_inicio = data_ref - timedelta(days=data_ref.weekday())
        dt_fim = dt_inicio + timedelta(days=6)
    except:
        messages.error(request, "Erro na data.")
        return redirect('area_gestor')

    equipe = get_object_or_404(Equipe, id=equipe_id)
    funcionarios = Funcionario.objects.filter(Q(equipe=equipe)|Q(outras_equipes=equipe)).distinct()

    # 2. Configurar PDF
    response = HttpResponse(content_type='application/pdf')
    nome_filial = equipe.nome.replace('Campo ', '').strip()
    filename = f"Pagamento_{nome_filial}_S{dt_inicio.isocalendar()[1]}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()

    # 3. Título
    elements.append(Paragraph(f"Relatório de Pagamento - {equipe.nome}", styles['Title']))
    elements.append(Paragraph(f"Período: {dt_inicio.strftime('%d/%m/%Y')} a {dt_fim.strftime('%d/%m/%Y')}", styles['Normal']))
    elements.append(Spacer(1, 20))

    # 4. Dados da Tabela (SEM COLUNA KM)
    # Colunas: Técnico | Banco | Ag/Conta/Op | Chave PIX | Valor Total
    data = [['Técnico', 'Banco', 'Ag/Conta/Op', 'Chave PIX', 'Valor Total (R$)']]
    
    total_geral = 0.0

    for func in funcionarios:
        kms = ControleKM.objects.filter(funcionario=func, data__range=[dt_inicio, dt_fim]).exclude(status='Rejeitado')
        despesas = DespesaDiversa.objects.filter(funcionario=func, data__range=[dt_inicio, dt_fim]).exclude(status='Rejeitado')
        
        total_km = sum([k.total_km for k in kms])
        total_despesas = sum([d.valor for d in despesas])

        # LÓGICA DE CÁLCULO CORRIGIDA (Padrão 1.20 se não tiver valor cadastrado)
        fator = float(func.valor_km) if func.valor_km and func.valor_km > 0 else 1.20
        valor_final = (float(total_km) * fator) + float(total_despesas)

        if valor_final > 0:
            dados_bancarios = f"{func.banco or '-'}"
            conta_info = f"Ag: {func.agencia or '-'} / CC: {func.conta or '-'} {func.operacao or ''}"
            
            data.append([
                func.nome_completo,
                dados_bancarios,
                conta_info,
                func.chave_pix or "-",
                f"R$ {valor_final:,.2f}"
            ])
            total_geral += valor_final

    # Linha Total
    data.append(['', '', '', 'TOTAL DA FILIAL:', f"R$ {total_geral:,.2f}"])

    # 5. Estilo da Tabela (Larguras Ajustadas para preencher o espaço sem a coluna KM)
    # A4 Landscape largura útil aprox ~780pts.
    # Distribuição: Técnico(230), Banco(120), Conta(170), Pix(160), Valor(100)
    table = Table(data, colWidths=[230, 120, 170, 160, 100])
    
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'), # Nomes à esquerda
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.beige), # Linha total
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    table.setStyle(style)
    elements.append(table)

    doc.build(elements)
    return response

@login_required
@require_POST
def atualizar_valor_km_equipe(request, equipe_id):
    equipe = get_object_or_404(Equipe, id=equipe_id)
    # Verifica permissão (Gestor da equipe ou RH)
    if not (request.user.is_superuser or usuario_eh_rh(request.user) or equipe.gestor == request.user.funcionario or equipe.gestores.filter(id=request.user.funcionario.id).exists()):
        messages.error(request, "Sem permissão para alterar valores desta equipe.")
        return redirect('area_gestor')

    novo_valor = request.POST.get('novo_valor_km')
    
    if novo_valor:
        try:
            # Converte virgula para ponto se necessário
            val = float(novo_valor.replace(',', '.'))
            
            # Atualiza TODOS os funcionários desta equipe (Principal e Secundária)
            funcs = Funcionario.objects.filter(Q(equipe=equipe)|Q(outras_equipes=equipe)).distinct()
            updated = funcs.update(valor_km=val)
            
            messages.success(request, f"Valor do KM atualizado para R$ {val:.2f} em {updated} colaboradores da filial.")
        except ValueError:
            messages.error(request, "Valor inválido inserido.")
    
    # Redireciona de volta mantendo os filtros
    return redirect(request.META.get('HTTP_REFERER', 'area_gestor'))

